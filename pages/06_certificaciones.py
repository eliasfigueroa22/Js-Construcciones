"""Tool 6 — Certificaciones y Avances de Obra.

Permite gestionar presupuestos por obra, crear certificaciones periódicas
de avance y generar documentos profesionales (PDF/Excel) para presentar
a clientes. Los datos se leen desde Supabase.
"""

from __future__ import annotations

import io
from datetime import date, datetime

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from connectors.supabase_connector import (
    clear_cache,
    create_record,
    delete_record,
    get_all_records,
    update_record,
)
from core.base_tool import ToolMetadata
from generators.pdf_generator import (
    generate_certificado_detalle_pdf,
    generate_certificado_memo_pdf,
    generate_resumen_financiero_pdf,
)

TOOL = ToolMetadata(
    name="Certificaciones",
    description="Certificados de avance de obra y gestión de presupuestos.",
    icon="📋",
    page_file="06_certificaciones.py",
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_UNIDAD_OPTIONS = ["m2", "m3", "glb", "ml", "Unid.", "bcs", "kg", "lt", "hs"]

_GRUPO_OPTIONS = [
    "Trabajos Previos", "Movimiento de Suelos", "Estructura",
    "Albañilería", "Revoques", "Impermeabilización", "Cubierta / Techo",
    "Cielorrasos", "Pisos y Zócalos", "Revestimientos", "Carpintería",
    "Herrería", "Pintura", "Instalación Eléctrica", "Instalación Sanitaria",
    "Instalación de Gas", "Aire Acondicionado", "Varios",
]

# ---------------------------------------------------------------------------
# CSS
# ---------------------------------------------------------------------------

_CSS = """
<style>
:root {
    --js-accent:  #E8622A;
    --js-success: #27AE60;
    --js-danger:  #E74C3C;
    --js-warn:    #E67E22;
    --js-muted:   #6B7280;
    --js-border:  rgba(255,255,255,0.09);
    --js-surface: rgba(255,255,255,0.03);
}
.js-sub  { color: var(--js-muted); font-size: .875rem; margin-top: -10px; margin-bottom: 24px; }
.js-pill { display: inline-block; border-radius: 3px; padding: 2px 8px; font-size: .68rem;
           font-weight: 700; letter-spacing: .8px; text-transform: uppercase; }
.cert-pill-borrador  { background: rgba(230,126,34,.18); color: #E67E22; }
.cert-pill-confirmado{ background: rgba(39,174,96,.18);  color: #27AE60; }
.cert-dirty  { color: var(--js-warn); font-size: .82rem; font-weight: 600; }
.cert-saved  { color: var(--js-muted); font-size: .82rem; }
.cert-sin-cotizar { background: rgba(231,76,60,.12); color: #E74C3C;
                    border-radius: 4px; padding: 4px 10px; font-size: .82rem; }
.cert-progress { height: 8px; border-radius: 4px; background: var(--js-border); overflow: hidden; }
.cert-progress-bar { height: 100%; border-radius: 4px; background: var(--js-accent); }
.cert-over { background: var(--js-danger) !important; }
</style>
"""

# ---------------------------------------------------------------------------
# Lookup helpers
# ---------------------------------------------------------------------------

def _fmt_fecha(iso: str) -> str:
    """Convierte 'YYYY-MM-DD' a 'DD/MM/YY'. Devuelve vacío si no hay fecha."""
    if not iso:
        return ""
    try:
        return datetime.strptime(iso[:10], "%Y-%m-%d").strftime("%d/%m/%y")
    except ValueError:
        return iso


def _fmt_gs(value) -> str:
    """Formatea monto en Gs.: sin decimales, miles con punto. Ej: Gs. 1.234.567"""
    try:
        return f"Gs. {int(float(value)):,}".replace(",", ".")
    except (ValueError, TypeError):
        return "Gs. 0"


def _fmt_gs_plain(value) -> str:
    """Igual que _fmt_gs pero sin prefijo 'Gs. ', para celdas de tabla."""
    try:
        return f"{int(float(value)):,}".replace(",", ".")
    except (ValueError, TypeError):
        return "0"


def _fmt_cant(value) -> str:
    """Formatea cantidad con 2 decimales, miles con punto, decimal con coma. Ej: 1.234,56"""
    try:
        v = float(value)
        formatted = f"{v:,.2f}"                              # "1,234.56"
        return formatted.replace(",", "X").replace(".", ",").replace("X", ".")  # "1.234,56"
    except (ValueError, TypeError):
        return "0,00"


def _fmt_pct(value) -> str:
    """Formatea porcentaje con 1 decimal y coma. Ej: 83,4%"""
    try:
        return f"{float(value):.1f}%".replace(".", ",")
    except (ValueError, TypeError):
        return "0,0%"


def _lineas_for_obra(obra_id, lineas: list[dict]) -> list[dict]:
    if not obra_id:
        return []
    result = [l for l in lineas if l.get("obra_id") == obra_id]
    result.sort(key=lambda x: int(x.get("orden") or 999))
    return result


def _ingresos_for_obra(obra_id, ingresos: list[dict]) -> list[dict]:
    if not obra_id:
        return []
    result = [i for i in ingresos if i.get("obra_id") == obra_id]
    result.sort(key=lambda x: str(x.get("fecha_ingreso") or ""))
    return result


def _cabs_for_obra(obra_id, cabeceras: list[dict]) -> list[dict]:
    if not obra_id:
        return []
    result = [c for c in cabeceras if c.get("obra_id") == obra_id]
    result.sort(key=lambda x: int(x.get("numero") or 0))
    return result


def _obra_has_confirmed_cert(obra_id, cabeceras: list[dict]) -> bool:
    return any(
        c.get("obra_id") == obra_id and c.get("estado") == "Confirmado"
        for c in cabeceras
    )


def _next_cert_numero(obra_id, cabeceras: list[dict]) -> int:
    nums = [
        int(c.get("numero") or 0)
        for c in cabeceras
        if c.get("obra_id") == obra_id
    ]
    return max(nums, default=0) + 1


# ---------------------------------------------------------------------------
# Computation engine
# ---------------------------------------------------------------------------

def _compute_anteriores(
    obra_id,
    cert_numero: int,
    cabeceras: list[dict],
    cert_lineas: list[dict],
) -> dict:
    """Sum cantidad_certificada from all Confirmado certs with numero < cert_numero."""
    cab_ids = {
        c["id"]
        for c in cabeceras
        if c.get("obra_id") == obra_id
        and c.get("estado") == "Confirmado"
        and int(c.get("numero") or 0) < cert_numero
    }
    anteriores: dict = {}
    for cl in cert_lineas:
        if cl.get("cabecera_id") in cab_ids:
            pl_id = cl.get("presupuesto_linea_id")
            anteriores[pl_id] = anteriores.get(pl_id, 0) + float(cl.get("cantidad_certificada") or 0)
    return anteriores


def _build_display_rows(
    presup_lineas: list[dict],
    cert_lineas_current: dict,
    anteriores: dict,
) -> list[dict]:
    """Build the full display table for a certification."""
    rows = []
    for pl in presup_lineas:
        pl_id   = pl["id"]
        cant_p  = float(pl.get("cantidad") or 0)
        pu      = float(pl.get("precio_unitario") or 0)
        pt      = cant_p * pu
        obs     = pl.get("observaciones") or ""

        cant_ant = anteriores.get(pl_id, 0)
        cant_act = cert_lineas_current.get(pl_id, 0)
        cant_acu = cant_ant + cant_act

        monto_ant = cant_ant * pu
        monto_act = cant_act * pu
        monto_acu = cant_acu * pu

        pct_ant = (cant_ant / cant_p * 100) if cant_p else 0
        pct_act = (cant_act / cant_p * 100) if cant_p else 0
        pct_acu = (cant_acu / cant_p * 100) if cant_p else 0

        rows.append({
            "id":           pl_id,
            "Zona":         pl.get("zona", ""),
            "ItemNro":      pl.get("item_nro", ""),
            "GrupoNombre":  pl.get("grupo_nombre", ""),
            "Rubro":        pl.get("rubro", ""),
            "Unidad":       pl.get("unidad", ""),
            "CantPres":     cant_p,
            "PU":           pu,
            "PTotal":       pt,
            "Observaciones": obs,
            "CantAnt":      cant_ant,
            "CantActual":   cant_act,
            "CantAcum":     cant_acu,
            "MontoAnt":     monto_ant,
            "MontoActual":  monto_act,
            "MontoAcum":    monto_acu,
            "PctAnt":       pct_ant,
            "PctActual":    pct_act,
            "PctAcum":      pct_acu,
            "CantRestante": cant_p - cant_ant,
            "SinCotizar":   bool(pl.get("sin_cotizar", False)),
        })
    return rows


# ---------------------------------------------------------------------------
# Session state
# ---------------------------------------------------------------------------

_SS_DEFAULTS = {
    # Presupuesto tab
    "cert_pres_obra_id":    None,
    "cert_pres_df":         None,      # pd.DataFrame for data_editor (source of truth)
    "cert_pres_ids":        "list",    # list[str|None] parallel to df rows (record IDs)
    "cert_pres_deleted_ids": "list",
    "cert_pres_dirty":      False,
    "cert_pres_saved_at":   None,
    "cert_pres_pending_delete": False,
    # Certificacion tab
    "cert_cab_id":          None,
    "cert_cab":             "cab",
    "cert_lineas_edit":     "dict",
    "cert_deleted_ids":     "list",
    "cert_dirty":           False,
    "cert_saved_at":        None,
    "cert_pending_delete":  False,
}


def _init_session_state():
    for key, default in _SS_DEFAULTS.items():
        if key not in st.session_state:
            if default == "list":
                st.session_state[key] = []
            elif default == "dict":
                st.session_state[key] = {}
            elif default == "cab":
                st.session_state[key] = {
                    "obra_id": None, "numero": 1,
                    "fecha_certificado": None, "estado": "Borrador",
                    "observaciones": "",
                }
            else:
                st.session_state[key] = default


def _build_pres_df(lineas: list[dict]) -> pd.DataFrame:
    """Build presupuesto DataFrame from Supabase records."""
    if not lineas:
        return pd.DataFrame(columns=[
            "Zona", "Grupo", "Item", "Rubro", "Unidad",
            "Cantidad", "P.U.", "Total", "S/C", "Obs.",
        ])
    return pd.DataFrame([
        {
            "Zona":     l.get("zona", ""),
            "Grupo":    l.get("grupo_nombre", ""),
            "Item":     l.get("item_nro", ""),
            "Rubro":    l.get("rubro", ""),
            "Unidad":   l.get("unidad", ""),
            "Cantidad": float(l.get("cantidad") or 0),
            "P.U.":     float(l.get("precio_unitario") or 0),
            "Total":    _fmt_gs_plain(float(l.get("cantidad") or 0) * float(l.get("precio_unitario") or 0)),
            "S/C":      bool(l.get("sin_cotizar", False)),
            "Obs.":     l.get("observaciones", ""),
        }
        for l in lineas
    ])


def _pres_df_to_lineas(df: pd.DataFrame, ids: list) -> list[dict]:
    """Convert edited DataFrame back to list of dicts for Supabase save."""
    lineas = []
    for i, row in df.iterrows():
        idx = int(i)
        rec_id = ids[idx] if idx < len(ids) else None
        lineas.append({
            "id":            rec_id,
            "zona":          str(row.get("Zona", "") or ""),
            "item_nro":      str(row.get("Item", "") or ""),
            "grupo_nombre":  str(row.get("Grupo", "") or ""),
            "rubro":         str(row.get("Rubro", "") or ""),
            "unidad":        str(row.get("Unidad", "") or ""),
            "cantidad":      float(row.get("Cantidad") or 0),
            "precio_unitario": float(row.get("P.U.") or 0),
            "observaciones": str(row.get("Obs.", "") or ""),
            "sin_cotizar":   bool(row.get("S/C", False)),
            "orden":         idx + 1,
        })
    return lineas


def _clear_caches():
    """Clear all cached data."""
    clear_cache()


def _reset_pres_state():
    st.session_state["cert_pres_obra_id"]       = None
    st.session_state["cert_pres_df"]            = None
    st.session_state["cert_pres_ids"]           = []
    st.session_state["cert_pres_deleted_ids"]   = []
    st.session_state["cert_pres_dirty"]         = False
    st.session_state["cert_pres_saved_at"]      = None
    st.session_state["cert_pres_pending_delete"] = False
    st.session_state.pop("pres_editor", None)


def _eliminar_presupuesto(obra_id) -> bool:
    """Elimina todas las op_cert_presupuesto_linea de una obra. Bloqueado si hay certs confirmadas."""
    ids = [rid for rid in st.session_state["cert_pres_ids"] if rid]
    for rec_id in ids:
        try:
            delete_record("op_cert_presupuesto_linea", rec_id)
        except Exception as e:
            st.error(f"Error al eliminar ítem: {e}")
            return False
    _reset_pres_state()
    _clear_caches()
    return True


def _reset_cert_state():
    st.session_state["cert_cab_id"]       = None
    st.session_state["cert_cab"]          = {
        "obra_id": None, "numero": 1,
        "fecha_certificado": None, "estado": "Borrador",
        "observaciones": "",
    }
    st.session_state["cert_lineas_edit"]  = {}
    st.session_state["cert_deleted_ids"]  = []
    st.session_state["cert_dirty"]        = False
    st.session_state["cert_saved_at"]     = None
    st.session_state["cert_pending_delete"] = False


def _on_pres_edit():
    """on_change callback for the presupuesto data_editor.

    Reads the edit delta from st.session_state["pres_editor"], applies it
    to the source DataFrame (cert_pres_df), and updates cert_pres_ids.
    Runs BEFORE the script body re-executes, so the updated df is ready
    for the next render cycle without triggering a second rerun.
    """
    changes = st.session_state.get("pres_editor")
    if not changes:
        return

    df = st.session_state["cert_pres_df"].copy()
    ids = list(st.session_state["cert_pres_ids"])

    # 1) Apply cell edits
    for row_idx_str, updates in changes.get("edited_rows", {}).items():
        row_idx = int(row_idx_str)
        for col, val in updates.items():
            df.at[row_idx, col] = val

    # 2) Append added rows (always append, even if the row is empty — dropping the
    #    "if new_row:" guard is critical: empty rows are {} which is falsy, so without
    #    this fix they're never stored in cert_pres_df → next rerun Streamlit sees a
    #    different-length DataFrame → widget resets → all pending edits vanish)
    for new_row in changes.get("added_rows", []):
        new_series = {
            "Zona": "", "Grupo": "", "Item": "", "Rubro": "", "Unidad": "",
            "Cantidad": 0.0, "P.U.": 0.0, "Total": 0.0, "S/C": False, "Obs.": "",
        }
        new_series.update(new_row)
        df = pd.concat([df, pd.DataFrame([new_series])], ignore_index=True)
        ids.append(None)

    # 3) Remove deleted rows
    for del_idx in sorted(changes.get("deleted_rows", []), reverse=True):
        if del_idx < len(ids):
            rec_id = ids[del_idx]
            if rec_id:
                st.session_state["cert_pres_deleted_ids"].append(rec_id)
            ids.pop(del_idx)
    if changes.get("deleted_rows"):
        df = df.drop(index=changes["deleted_rows"]).reset_index(drop=True)

    # 4) Recompute Total column (pre-formatted string)
    df["Total"] = (
        df["Cantidad"].fillna(0).astype(float) * df["P.U."].fillna(0).astype(float)
    ).apply(_fmt_gs_plain)

    st.session_state["cert_pres_df"]  = df
    st.session_state["cert_pres_ids"] = ids
    st.session_state["cert_pres_dirty"] = True


# ---------------------------------------------------------------------------
# Cascading obra filter (shared by tabs)
# ---------------------------------------------------------------------------

def _render_obra_selector(
    prefix: str,
    obras_raw: list[dict],
    disabled: bool = False,
    allow_all: bool = False,
) -> int | None:
    """Render cascading Estado → Categoría → Obra selector. Returns obra_id or None.

    If allow_all=True, adds a "Todas las obras" option that returns None.
    """
    estados = sorted({r.get("estado_obra", "") for r in obras_raw if r.get("estado_obra")})
    c1, c2, c3 = st.columns([1, 1, 2])

    with c1:
        estado_sel = st.selectbox(
            "Estado de obra", ["Todos"] + estados,
            key=f"{prefix}_filtro_estado", disabled=disabled,
        )
    filtradas = [
        r for r in obras_raw
        if estado_sel == "Todos" or r.get("estado_obra") == estado_sel
    ]

    with c2:
        cats = sorted({r.get("categoria_obra", "") for r in filtradas if r.get("categoria_obra")})
        cat_sel = st.selectbox(
            "Categoría", ["Todas"] + cats,
            key=f"{prefix}_filtro_cat", disabled=disabled,
        )
    filtradas = [
        r for r in filtradas
        if cat_sel == "Todas" or r.get("categoria_obra") == cat_sel
    ]

    with c3:
        obra_opts_map = {}
        for r in sorted(filtradas, key=lambda x: x.get("clave", "")):
            clave = r.get("clave", r.get("nombre", ""))
            obra_opts_map[clave] = r["id"]
        if not obra_opts_map:
            st.info("No hay obras con los filtros seleccionados.")
            return None
        options = (["Todas las obras"] if allow_all else []) + list(obra_opts_map.keys())
        obra_sel = st.selectbox(
            "Obra", options,
            key=f"{prefix}_filtro_obra", disabled=disabled,
        )
        if allow_all and obra_sel == "Todas las obras":
            return None
    return obra_opts_map.get(obra_sel)


# ---------------------------------------------------------------------------
# Presupuesto — Supabase write operations
# ---------------------------------------------------------------------------

def _guardar_presupuesto(obra_id) -> bool:
    # cert_pres_df already has all edits applied via the on_change callback
    source_df = st.session_state.get("cert_pres_df")
    if source_df is None or source_df.empty:
        st.warning("No hay ítems para guardar.")
        return False
    ids = st.session_state["cert_pres_ids"]
    lineas = _pres_df_to_lineas(source_df, ids)

    # Phase 1: delete removed lines
    for rec_id in st.session_state["cert_pres_deleted_ids"]:
        try:
            delete_record("op_cert_presupuesto_linea", rec_id)
        except Exception as e:
            st.error(f"Error al eliminar línea: {e}")
            return False

    # Phase 2: upsert lines
    new_ids = []
    for idx, linea in enumerate(lineas):
        fields = {
            "obra_id":          obra_id,
            "orden":            idx + 1,
            "zona":             linea.get("zona", ""),
            "item_nro":         linea.get("item_nro", ""),
            "grupo_nombre":     linea.get("grupo_nombre", ""),
            "rubro":            linea.get("rubro", ""),
            "unidad":           linea.get("unidad", ""),
            "cantidad":         linea.get("cantidad") or None,
            "precio_unitario":  linea.get("precio_unitario") or None,
            "observaciones":    linea.get("observaciones", ""),
            "sin_cotizar":      bool(linea.get("sin_cotizar", False)),
        }
        try:
            if linea.get("id"):
                update_record("op_cert_presupuesto_linea", linea["id"], fields)
                new_ids.append(linea["id"])
            else:
                new_id = create_record("op_cert_presupuesto_linea", fields)
                new_ids.append(new_id)
        except Exception as e:
            st.error(f"Error al guardar línea {idx + 1}: {e}")
            return False

    st.session_state["cert_pres_ids"]        = new_ids
    st.session_state["cert_pres_deleted_ids"] = []
    st.session_state["cert_pres_dirty"]       = False
    st.session_state["cert_pres_saved_at"]    = datetime.now()
    st.session_state.pop("pres_editor", None)
    _clear_caches()
    return True


# ---------------------------------------------------------------------------
# Certificacion — Supabase write operations
# ---------------------------------------------------------------------------

def _guardar_cert(obra_id, presup_lineas: list[dict]) -> bool:
    cab = st.session_state["cert_cab"]

    fecha_val = cab.get("fecha_certificado")
    if isinstance(fecha_val, date):
        fecha_str = fecha_val.isoformat()
    elif isinstance(fecha_val, str):
        fecha_str = fecha_val[:10]
    else:
        fecha_str = ""

    # Build cert_ref
    obras_raw = get_all_records("dim_obra")
    obra_rec = next((o for o in obras_raw if o["id"] == obra_id), {})
    obra_clave = obra_rec.get("clave", "OBRA")
    cert_ref = f"{obra_clave}-CERT-{cab.get('numero', 1)}"

    cab_fields = {
        "cert_ref":          cert_ref,
        "obra_id":           obra_id,
        "numero":            cab.get("numero", 1),
        "fecha_certificado": fecha_str,
        "estado":            cab.get("estado", "Borrador"),
        "observaciones":     cab.get("observaciones", ""),
    }

    try:
        if st.session_state["cert_cab_id"]:
            update_record("op_cert_cabecera", st.session_state["cert_cab_id"], cab_fields)
        else:
            # Guard: check if a record with same obra_id+numero already exists in Supabase
            # to avoid creating duplicates when cert_cab_id was lost from session state
            all_cabs = get_all_records("op_cert_cabecera")
            existing = next(
                (
                    c for c in all_cabs
                    if c.get("obra_id") == obra_id
                    and int(c.get("numero") or 0) == int(cab.get("numero", 1))
                ),
                None,
            )
            if existing:
                update_record("op_cert_cabecera", existing["id"], cab_fields)
                st.session_state["cert_cab_id"] = existing["id"]
            else:
                new_id = create_record("op_cert_cabecera", cab_fields)
                st.session_state["cert_cab_id"] = new_id
    except Exception as e:
        st.error(f"Error al guardar cabecera: {e}")
        return False

    cab_id = st.session_state["cert_cab_id"]
    cert_lineas_edit = st.session_state["cert_lineas_edit"]

    # Phase: delete removed cert lines
    for rec_id in st.session_state["cert_deleted_ids"]:
        try:
            delete_record("op_cert_linea", rec_id)
        except Exception as e:
            st.error(f"Error al eliminar línea cert: {e}")
            return False

    # Phase: upsert cert lines (only for presup lines that have a non-zero cantidad_certificada)
    all_cert_lineas = get_all_records("op_cert_linea")
    existing_map = {}
    for cl in all_cert_lineas:
        if cl.get("cabecera_id") == cab_id:
            pl_id = cl.get("presupuesto_linea_id")
            existing_map[pl_id] = cl["id"]

    for pl in presup_lineas:
        pl_id = pl["id"]
        cant = cert_lineas_edit.get(pl_id, 0)
        fields = {
            "cabecera_id":          cab_id,
            "presupuesto_linea_id": pl_id,
            "cantidad_certificada": cant if cant else None,
            "linea_ref":            f"{cert_ref}-{pl.get('item_nro', '')}",
        }
        try:
            if pl_id in existing_map:
                update_record("op_cert_linea", existing_map[pl_id], fields)
            else:
                if cant:
                    create_record("op_cert_linea", fields)
        except Exception as e:
            st.error(f"Error al guardar línea cert: {e}")
            return False

    st.session_state["cert_deleted_ids"] = []
    st.session_state["cert_dirty"]       = False
    st.session_state["cert_saved_at"]    = datetime.now()
    _clear_caches()
    return True


def _eliminar_cert(cert_lineas_all: list[dict]) -> bool:
    """Elimina la cert en sesión: primero sus op_cert_lineas, luego la op_cert_cabecera."""
    cab_id = st.session_state["cert_cab_id"]
    if not cab_id:
        return False
    lineas = [cl for cl in cert_lineas_all if cl.get("cabecera_id") == cab_id]
    for cl in lineas:
        try:
            delete_record("op_cert_linea", cl["id"])
        except Exception as e:
            st.error(f"Error al eliminar línea de certificación: {e}")
            return False
    try:
        delete_record("op_cert_cabecera", cab_id)
    except Exception as e:
        st.error(f"Error al eliminar la cabecera: {e}")
        return False
    _reset_cert_state()
    _clear_caches()
    return True


def _confirmar_cert(obra_id, presup_lineas: list[dict]) -> bool:
    if not _guardar_cert(obra_id, presup_lineas):
        return False
    try:
        update_record("op_cert_cabecera", st.session_state["cert_cab_id"], {"estado": "Confirmado"})
    except Exception as e:
        st.error(f"Error al confirmar: {e}")
        return False
    st.session_state["cert_cab"]["estado"] = "Confirmado"
    st.session_state["cert_dirty"] = False
    _clear_caches()
    return True


def _reabrir_cert() -> bool:
    try:
        update_record("op_cert_cabecera", st.session_state["cert_cab_id"], {"estado": "Borrador"})
    except Exception as e:
        st.error(f"Error al reabrir: {e}")
        return False
    st.session_state["cert_cab"]["estado"] = "Borrador"
    _clear_caches()
    return True


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def _validate_cert(cab: dict, cert_lineas_edit: dict) -> list[str]:
    errors = []
    if not cab.get("fecha_certificado"):
        errors.append("Ingresá la fecha del certificado.")
    return errors


def _check_overcertification(display_rows: list[dict]) -> list[str]:
    warnings = []
    for r in display_rows:
        if r.get("SinCotizar"):
            continue
        if r["CantActual"] > 0 and r["CantAcum"] > r["CantPres"] and r["CantPres"] > 0:
            warnings.append(
                f"Ítem {r['ItemNro']} ({r['Rubro'][:40]}): "
                f"acumulado {_fmt_cant(r['CantAcum'])} supera presupuesto {_fmt_cant(r['CantPres'])}"
            )
    return warnings


# ---------------------------------------------------------------------------
# Excel import for presupuesto
# ---------------------------------------------------------------------------

def _parse_presupuesto_excel(file) -> list[dict] | None:
    """Parse an Excel file into presupuesto lines. Returns list of dicts or None on error."""
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        st.error(f"Error al leer el archivo Excel: {e}")
        return None

    lineas = []
    current_grupo = ""
    orden = 1

    for row in ws.iter_rows(min_row=2, values_only=False):
        cells = [c.value for c in row]
        # Skip empty rows
        if not any(cells):
            continue

        # Try to detect: Item | Rubro | Unid | Cant | PU
        item_val  = cells[0]
        rubro_val = cells[1] if len(cells) > 1 else ""
        unid_val  = cells[2] if len(cells) > 2 else ""
        cant_val  = cells[3] if len(cells) > 3 else 0
        pu_val    = cells[4] if len(cells) > 4 else 0

        if not rubro_val:
            continue

        rubro_str = str(rubro_val).strip()
        item_str  = str(item_val).strip() if item_val else ""
        unid_str  = str(unid_val).strip() if unid_val else ""

        # Detect group headers: rows where there is no unit/quantity
        obs = ""
        if len(cells) > 5 and cells[5]:
            obs = str(cells[5]).strip()

        try:
            cant_f = float(cant_val) if cant_val else 0
        except (ValueError, TypeError):
            cant_f = 0
        try:
            pu_f = float(pu_val) if pu_val else 0
        except (ValueError, TypeError):
            pu_f = 0

        # Detect if this is a group header (has item number, rubro ends with ":", no unit)
        is_group = rubro_str.endswith(":") and not unid_str
        if is_group:
            current_grupo = rubro_str.rstrip(":")
            continue

        # Detect sub-items (item like "6a", "6b")
        grupo = ""
        if item_str and any(c.isalpha() for c in item_str):
            grupo = current_grupo

        lineas.append({
            "id":            None,
            "item_nro":      item_str,
            "grupo_nombre":  grupo,
            "rubro":         rubro_str,
            "unidad":        unid_str,
            "cantidad":      cant_f,
            "precio_unitario": pu_f,
            "observaciones": obs,
            "orden":         orden,
        })
        orden += 1

    return lineas if lineas else None


# ---------------------------------------------------------------------------
# Excel export for certification
# ---------------------------------------------------------------------------

def _generate_cert_excel(
    obra_nombre: str,
    cert_numero: int,
    fecha_cert: str,
    display_rows: list[dict],
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Certificado {cert_numero}"

    accent = "E8622A"
    dark   = "1A252F"
    gray   = "F2F3F4"

    hdr_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    hdr_fill   = PatternFill("solid", fgColor=accent)
    title_font = Font(name="Calibri", bold=True, color=dark, size=12)
    meta_font  = Font(name="Calibri", color="555555", size=9)
    data_font  = Font(name="Calibri", size=9)
    total_font = Font(name="Calibri", bold=True, size=9)
    total_fill = PatternFill("solid", fgColor="D5D8DC")
    alt_fill   = PatternFill("solid", fgColor=gray)
    thin_side  = Side(style="thin", color="CCCCCC")
    thin_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    # Title
    ws.merge_cells("A1:N1")
    ws["A1"] = f"PLANILLA DE CERTIFICADO Nº {cert_numero}"
    ws["A1"].font = title_font

    # Metadata
    ws.merge_cells("A2:N2")
    ws["A2"] = f"Obra: {obra_nombre}  |  Fecha certificado: {fecha_cert}"
    ws["A2"].font = meta_font

    # Headers
    headers = [
        "Item", "Rubro", "Unid.", "Cant.Pres.", "P.U.", "P.Total",
        "Cant.Ant.", "Cant.Actual", "Cant.Acum.",
        "M.Anterior", "M.Actual", "M.Acumulado",
        "% Ant.", "% Acum.",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    current_grupo = ""
    row_num = 5
    for i, r in enumerate(display_rows):
        # Group header
        grupo = r.get("GrupoNombre", "")
        if grupo and grupo != current_grupo:
            current_grupo = grupo
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=14)
            cell = ws.cell(row=row_num, column=1, value=grupo)
            cell.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5D6D7E")
            cell.border = thin_border
            row_num += 1

        fill = alt_fill if i % 2 == 0 else PatternFill()
        values = [
            r["ItemNro"], r["Rubro"], r["Unidad"], r["CantPres"], r["PU"], r["PTotal"],
            r["CantAnt"], r["CantActual"], r["CantAcum"],
            r["MontoAnt"], r["MontoActual"], r["MontoAcum"],
            r["PctAnt"], r["PctAcum"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.border = thin_border
            if col_idx >= 4:
                cell.alignment = Alignment(horizontal="right")
                if col_idx in (5, 6, 10, 11, 12):
                    cell.number_format = "#,##0"
                elif col_idx in (4, 7, 8, 9):
                    cell.number_format = "#,##0.000"
                elif col_idx in (13, 14):
                    cell.number_format = "0.00%"
                    cell.value = val / 100 if val else 0
        row_num += 1

    # Totals row
    total_pres  = sum(r["PTotal"] for r in display_rows)
    total_m_ant = sum(r["MontoAnt"] for r in display_rows)
    total_m_act = sum(r["MontoActual"] for r in display_rows)
    total_m_acu = sum(r["MontoAcum"] for r in display_rows)
    pct_ant_tot = (total_m_ant / total_pres * 100) if total_pres else 0
    pct_acu_tot = (total_m_acu / total_pres * 100) if total_pres else 0

    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
    cell = ws.cell(row=row_num, column=1, value="TOTAL GS SIN IVA")
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border

    totals = [None, None, None, None, None, total_pres, None, None, None,
              total_m_ant, total_m_act, total_m_acu, pct_ant_tot, pct_acu_tot]
    for col_idx, val in enumerate(totals, 1):
        if col_idx <= 5:
            continue
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="right")
        if col_idx in (6, 10, 11, 12):
            cell.number_format = "#,##0"
        elif col_idx in (13, 14):
            cell.number_format = "0.00%"
            cell.value = val / 100 if val else 0

    # Column widths
    widths = [6, 45, 6, 10, 12, 14, 10, 10, 10, 14, 14, 14, 8, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:N{row_num}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Tab 1 — Presupuesto
# ---------------------------------------------------------------------------

def _render_tab_presupuesto(obras_raw: list[dict]):
    st.markdown(
        '<p class="js-sub">Definí los ítems del presupuesto para cada obra. '
        'Este es el contrato base contra el cual se certifican los avances.</p>',
        unsafe_allow_html=True,
    )

    cabeceras_raw = get_all_records("op_cert_cabecera")
    all_presup    = get_all_records("op_cert_presupuesto_linea")

    obra_id = _render_obra_selector("pres", obras_raw)
    if not obra_id:
        return

    is_locked = _obra_has_confirmed_cert(obra_id, cabeceras_raw)
    presup_lineas = _lineas_for_obra(obra_id, all_presup)

    if is_locked:
        st.warning("Este presupuesto no puede modificarse porque tiene certificaciones confirmadas.")

    # Check if we need to load into state (new obra or obra changed)
    if st.session_state["cert_pres_obra_id"] != obra_id:
        st.session_state["cert_pres_obra_id"] = obra_id
        st.session_state["cert_pres_df"]  = _build_pres_df(presup_lineas)
        st.session_state["cert_pres_ids"] = [l["id"] for l in presup_lineas]
        st.session_state["cert_pres_deleted_ids"] = []
        st.session_state["cert_pres_dirty"]       = False
        st.session_state["cert_pres_saved_at"]    = None
        st.session_state.pop("pres_editor", None)

    pres_df = st.session_state["cert_pres_df"]

    # --- Empty state: offer create / import / copy ---
    if (pres_df is None or pres_df.empty) and not is_locked:
        st.divider()
        st.markdown("##### Esta obra no tiene presupuesto aún")
        c1, c2, c3 = st.columns(3)
        with c1:
            if st.button("➕ Crear presupuesto vacío", use_container_width=True):
                st.session_state["cert_pres_df"] = _build_pres_df([{
                    "item_nro": "1", "grupo_nombre": "", "rubro": "",
                    "unidad": "m2", "cantidad": 0, "precio_unitario": 0,
                    "observaciones": "",
                }])
                st.session_state["cert_pres_ids"] = [None]
                st.session_state["cert_pres_dirty"] = True
                st.session_state.pop("pres_editor", None)
                st.rerun()
        with c2:
            uploaded = st.file_uploader(
                "📤 Importar Excel", type=["xlsx"],
                key="pres_upload", label_visibility="collapsed",
            )
            if uploaded:
                parsed = _parse_presupuesto_excel(uploaded)
                if parsed:
                    st.session_state["cert_pres_df"] = _build_pres_df(parsed)
                    st.session_state["cert_pres_ids"] = [l.get("id") for l in parsed]
                    st.session_state["cert_pres_dirty"] = True
                    st.session_state.pop("pres_editor", None)
                    st.rerun()
                else:
                    st.error("No se pudieron extraer ítems del archivo.")
        with c3:
            # Copy from another obra
            other_obras = [
                o for o in obras_raw
                if o["id"] != obra_id and _lineas_for_obra(o["id"], all_presup)
            ]
            if other_obras:
                copy_opts = {
                    o.get("clave", o.get("nombre", "")): o["id"]
                    for o in other_obras
                }
                sel_copy = st.selectbox(
                    "Copiar de", list(copy_opts.keys()),
                    key="pres_copy_obra", label_visibility="collapsed",
                )
                if st.button("📋 Copiar presupuesto", use_container_width=True):
                    source_lineas = _lineas_for_obra(copy_opts[sel_copy], all_presup)
                    st.session_state["cert_pres_df"] = _build_pres_df(source_lineas)
                    st.session_state["cert_pres_ids"] = [None] * len(source_lineas)
                    st.session_state["cert_pres_dirty"] = True
                    st.session_state.pop("pres_editor", None)
                    st.rerun()
        return

    # --- Presupuesto editor ---
    st.divider()

    # Import Excel (also available when lines exist)
    if not is_locked:
        with st.expander("📤 Importar desde Excel (reemplaza ítems actuales)"):
            uploaded = st.file_uploader(
                "Seleccionar archivo", type=["xlsx"],
                key="pres_upload_replace",
            )
            if uploaded:
                parsed = _parse_presupuesto_excel(uploaded)
                if parsed:
                    # Mark old lines for deletion
                    for rid in st.session_state["cert_pres_ids"]:
                        if rid and rid not in st.session_state["cert_pres_deleted_ids"]:
                            st.session_state["cert_pres_deleted_ids"].append(rid)
                    st.session_state["cert_pres_df"] = _build_pres_df(parsed)
                    st.session_state["cert_pres_ids"] = [l.get("id") for l in parsed]
                    st.session_state["cert_pres_dirty"] = True
                    st.session_state.pop("pres_editor", None)
                    st.rerun()

    # Display/edit the presupuesto table
    if is_locked:
        pres_df_display = pres_df.copy()
        pres_df_display["Cantidad"] = pres_df_display["Cantidad"].apply(_fmt_cant)
        pres_df_display["P.U."]     = pres_df_display["P.U."].apply(_fmt_gs_plain)
        st.dataframe(
            pres_df_display, use_container_width=True, hide_index=True,
            column_config={
                "Zona":    st.column_config.TextColumn(),
                "Grupo":   st.column_config.TextColumn(),
                "Cantidad":st.column_config.TextColumn(),
                "P.U.":    st.column_config.TextColumn("P.U. (Gs.)"),
                "Total":   st.column_config.TextColumn("Total (Gs.)"),
            },
        )
    else:
        st.data_editor(
            pres_df, use_container_width=True, hide_index=True,
            num_rows="dynamic",
            column_config={
                "Zona":      st.column_config.TextColumn(width="medium", help="Zona o bloque de la obra (ej: OFICINAS, COMEDOR). Dejar vacío si la obra no tiene zonas."),
                "Grupo":     st.column_config.SelectboxColumn(options=_GRUPO_OPTIONS, width="medium", help="Grupo de trabajo dentro de la zona."),
                "Item":      st.column_config.TextColumn(width="small"),
                "Rubro":     st.column_config.TextColumn(width="large"),
                "Unidad":    st.column_config.SelectboxColumn(options=_UNIDAD_OPTIONS, width="small"),
                "Cantidad":  st.column_config.NumberColumn(format="%.2f", width="small"),
                "P.U.":      st.column_config.NumberColumn(format=",.0f", width="small"),
                "Total":     st.column_config.TextColumn("Total (Gs.)", disabled=True, width="small"),
                "S/C":       st.column_config.CheckboxColumn(label="S/C", help="Sin Cotizar", width="small"),
                "Obs.":      st.column_config.TextColumn(width="small"),
            },
            key="pres_editor",
            on_change=_on_pres_edit,
        )

    # Summary metrics — pres_df already has edits applied via on_change callback
    total_items = len(pres_df)
    total_pres = (pres_df["Cantidad"].fillna(0) * pres_df["P.U."].fillna(0)).sum()
    mc1, mc2, mc3 = st.columns(3)
    mc1.metric("Ítems", total_items)
    mc2.metric("Total presupuesto", _fmt_gs(total_pres))
    with mc3:
        if st.session_state["cert_pres_dirty"]:
            st.markdown('<span class="cert-dirty">● Cambios sin guardar</span>', unsafe_allow_html=True)
        elif st.session_state["cert_pres_saved_at"]:
            ts = st.session_state["cert_pres_saved_at"].strftime("%H:%M:%S")
            st.markdown(f'<span class="cert-saved">Guardado a las {ts}</span>', unsafe_allow_html=True)

    # Action buttons
    if not is_locked:
        st.divider()
        c1, c2, c3 = st.columns([1, 1, 4])
        with c1:
            if st.button("💾 Guardar presupuesto", use_container_width=True, type="primary"):
                if _guardar_presupuesto(obra_id):
                    st.success("Presupuesto guardado.")
                    st.rerun()
        with c2:
            if st.button("🗑️ Eliminar presupuesto", use_container_width=True):
                st.session_state["cert_pres_pending_delete"] = True
                st.rerun()

    # ── Confirmation zone ────────────────────────────────────────────────────
    if not is_locked and st.session_state.get("cert_pres_pending_delete"):
        n_items = len([rid for rid in st.session_state["cert_pres_ids"] if rid])
        st.divider()
        st.error(
            f"**Estás por eliminar el presupuesto completo de esta obra** de forma permanente.\n\n"
            f"Esto borrará **{n_items} ítem(s)** de Airtable. "
            f"Esta acción **no se puede deshacer**."
        )
        confirmado = st.checkbox(
            "Entiendo que esta acción es irreversible y quiero continuar.",
            key="pres_delete_confirm_chk",
        )
        dc1, dc2, _ = st.columns([1, 1, 4])
        with dc1:
            if st.button(
                "Eliminar definitivamente",
                type="primary",
                disabled=not confirmado,
                use_container_width=True,
                key="pres_delete_exec",
            ):
                if _eliminar_presupuesto(obra_id):
                    st.success("Presupuesto eliminado correctamente.")
                    st.rerun()
        with dc2:
            if st.button("Cancelar", use_container_width=True, key="pres_delete_cancel"):
                st.session_state["cert_pres_pending_delete"] = False
                st.rerun()


# ---------------------------------------------------------------------------
# Tab 2 — Certificaciones
# ---------------------------------------------------------------------------

def _load_cert_into_state(cab: dict, cert_lineas_all: list[dict]):
    st.session_state["cert_cab_id"] = cab["id"]
    st.session_state["cert_cab"] = {
        "obra_id":          cab.get("obra_id"),
        "numero":           int(cab.get("numero") or 1),
        "fecha_certificado": cab.get("fecha_certificado"),
        "estado":           cab.get("estado", "Borrador"),
        "observaciones":    cab.get("observaciones", ""),
    }
    # Load cantidad_certificada by presupuesto_linea_id
    edit_map = {}
    for cl in cert_lineas_all:
        if cl.get("cabecera_id") == cab["id"]:
            pl_id = cl.get("presupuesto_linea_id")
            edit_map[pl_id] = float(cl.get("cantidad_certificada") or 0)
    st.session_state["cert_lineas_edit"] = edit_map
    st.session_state["cert_deleted_ids"] = []
    st.session_state["cert_dirty"]       = False
    st.session_state["cert_saved_at"]    = None


def _render_tab_certificaciones(obras_raw: list[dict], maps: dict):
    st.markdown(
        '<p class="js-sub">Creá certificaciones periódicas de avance contra el presupuesto de cada obra.</p>',
        unsafe_allow_html=True,
    )

    all_presup      = get_all_records("op_cert_presupuesto_linea")
    cabeceras_raw   = get_all_records("op_cert_cabecera")
    cert_lineas_all = get_all_records("op_cert_linea")

    # Only show obras that have at least one presupuesto line
    obras_ids_con_presup = {l.get("obra_id") for l in all_presup if l.get("obra_id")}
    obras_con_presup = [o for o in obras_raw if o["id"] in obras_ids_con_presup]

    if not obras_con_presup:
        st.info("No hay obras con presupuesto definido. Creá uno en la pestaña Presupuesto.")
        return

    obra_id = _render_obra_selector("cert", obras_con_presup)
    if not obra_id:
        return

    presup_lineas = _lineas_for_obra(obra_id, all_presup)
    if not presup_lineas:
        st.info("Esta obra no tiene presupuesto. Crealo primero en la pestaña Presupuesto.")
        return

    obra_cabs = _cabs_for_obra(obra_id, cabeceras_raw)

    # --- Load/New section ---
    st.divider()
    c_sel, c_btn_load, c_btn_nueva = st.columns([3, 1, 1])

    cab_opts: dict[str, int] = {}
    sel_label = None
    with c_sel:
        if obra_cabs:
            for c in obra_cabs:
                num   = c.get("numero", "?")
                fecha = str(c.get("fecha_certificado", ""))[:10]
                est   = c.get("estado", "")
                label = f"Cert. Nº {num} — {fecha} [{est}]"
                cab_opts[label] = c["id"]
            sel_label = st.selectbox(
                f"Certificaciones ({len(obra_cabs)})",
                list(cab_opts.keys()), key="cert_sel_cab",
            )
        else:
            st.info("No hay certificaciones para esta obra.")

    with c_btn_load:
        st.markdown("<br>", unsafe_allow_html=True)
        if sel_label and st.button("📂 Cargar", key="cert_btn_load", use_container_width=True):
            sel_id  = cab_opts[sel_label]
            sel_cab = next(c for c in cabeceras_raw if c["id"] == sel_id)
            _load_cert_into_state(sel_cab, cert_lineas_all)
            st.rerun()

    with c_btn_nueva:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("➕ Nueva", key="cert_btn_nueva", use_container_width=True):
            next_num = _next_cert_numero(obra_id, cabeceras_raw)
            _reset_cert_state()
            st.session_state["cert_cab"]["obra_id"] = obra_id
            st.session_state["cert_cab"]["numero"] = next_num
            st.session_state["cert_cab"]["fecha_certificado"] = date.today()
            st.rerun()

    # If a cert is already loaded for this obra, show it without requiring "Cargar" again
    # (handles the case where the user saved a new cert and the page reruns)
    cab = st.session_state["cert_cab"]
    if (
        not cab.get("obra_id")
        and sel_label
        and st.session_state["cert_cab_id"]
    ):
        sel_id = cab_opts.get(sel_label)
        if sel_id == st.session_state["cert_cab_id"]:
            sel_cab = next((c for c in cabeceras_raw if c["id"] == sel_id), None)
            if sel_cab:
                _load_cert_into_state(sel_cab, cert_lineas_all)

    # --- Editing section ---
    cab = st.session_state["cert_cab"]
    if not cab.get("obra_id"):
        return

    # Make sure we're looking at the right obra
    if cab["obra_id"] != obra_id:
        _reset_cert_state()
        return

    locked = cab.get("estado") == "Confirmado"

    st.divider()

    # Status pill
    estado = cab.get("estado", "Borrador")
    pill_cls = "cert-pill-confirmado" if estado == "Confirmado" else "cert-pill-borrador"
    st.markdown(
        f'##### Certificado Nº {cab.get("numero", "?")} '
        f'<span class="js-pill {pill_cls}">{estado}</span>',
        unsafe_allow_html=True,
    )

    # Header form
    hc1, hc2 = st.columns([1, 3])
    with hc1:
        fecha_val = cab.get("fecha_certificado")
        if isinstance(fecha_val, str) and fecha_val:
            try:
                fecha_val = datetime.strptime(fecha_val[:10], "%Y-%m-%d").date()
            except ValueError:
                fecha_val = date.today()
        elif not fecha_val:
            fecha_val = date.today()
        new_fecha = st.date_input("Fecha del certificado", value=fecha_val, disabled=locked, key="cert_fecha")
        if new_fecha != cab.get("fecha_certificado"):
            cab["fecha_certificado"] = new_fecha
            st.session_state["cert_dirty"] = True
    with hc2:
        new_obs = st.text_input("Observaciones", value=cab.get("observaciones", ""), disabled=locked, key="cert_obs")
        if new_obs != cab.get("observaciones", ""):
            cab["observaciones"] = new_obs
            st.session_state["cert_dirty"] = True

    # Compute display rows
    cert_numero    = cab.get("numero", 1)
    anteriores     = _compute_anteriores(obra_id, cert_numero, cabeceras_raw, cert_lineas_all)
    cert_lineas_ed = st.session_state["cert_lineas_edit"]
    display_rows   = _build_display_rows(presup_lineas, cert_lineas_ed, anteriores)

    # ── 3-level hierarchy: Zona → Grupo → Rubro ─────────────────────────────────
    # Build zona_map: {zona: {grupo: [rows]}} preserving insertion order.
    zona_map: dict[str, dict[str, list[dict]]] = {}
    for r in display_rows:
        z = r.get("Zona") or ""
        g = r.get("GrupoNombre") or ""
        if z not in zona_map:
            zona_map[z] = {}
        if g not in zona_map[z]:
            zona_map[z][g] = []
        zona_map[z][g].append(r)

    _COL_CFG: dict = {
        "Item":        st.column_config.TextColumn(width="small"),
        "Rubro":       st.column_config.TextColumn(width="large"),
        "Unid.":       st.column_config.TextColumn(width="small"),
        "Cant.Pres.":  st.column_config.TextColumn(width="small"),
        "P.U.":        st.column_config.TextColumn("P.U. (Gs.)", width="small"),
        "P.Total":     st.column_config.TextColumn("P.Total (Gs.)", width="small"),
        "Cant.Ant.":   st.column_config.TextColumn(width="small"),
        "Cant.Actual": st.column_config.NumberColumn(format="%.2f", step=0.01, width="small"),
        "Cant.Acum.":  st.column_config.TextColumn(width="small"),
        "M.Ant.":      st.column_config.TextColumn("M.Ant. (Gs.)", width="small"),
        "M.Actual":    st.column_config.TextColumn("M.Actual (Gs.)", width="small"),
        "M.Acum.":     st.column_config.TextColumn("M.Acum. (Gs.)", width="small"),
        "% Acum.":     st.column_config.ProgressColumn(min_value=0, max_value=100, format="%.1f%%"),
    }
    _DISABLED = [
        "Item", "Rubro", "Unid.", "Cant.Pres.", "P.U.", "P.Total",
        "Cant.Ant.", "Cant.Acum.", "M.Ant.", "M.Actual", "M.Acum.", "% Acum.",
    ]

    new_edit: dict = {}
    editor_idx = 0
    for zona_name, grupos in zona_map.items():
        # Zona header (level 1) — only rendered if the obra has zones
        if zona_name:
            st.markdown(
                f'<div style="background:rgba(30,50,80,.95);color:#fff;'
                f'padding:7px 14px;font-weight:800;border-radius:5px;'
                f'margin:16px 0 4px 0;font-size:.9rem;letter-spacing:.8px;'
                f'border-left:4px solid #E8622A">'
                f'{zona_name.upper()}</div>',
                unsafe_allow_html=True,
            )

        for grupo_name, g_rows in grupos.items():
            # Grupo header (level 2) — only rendered if group has a name
            if grupo_name:
                st.markdown(
                    f'<div style="background:rgba(61,90,128,.75);color:#fff;'
                    f'padding:4px 12px 4px 20px;font-weight:700;border-radius:4px;'
                    f'margin:6px 0 2px 0;font-size:.78rem;letter-spacing:.6px">'
                    f'{grupo_name.upper()}</div>',
                    unsafe_allow_html=True,
                )

            df_section = pd.DataFrame([
                {
                    "Item":        r["ItemNro"],
                    "Rubro":       r["Rubro"],
                    "Unid.":       r["Unidad"],
                    "Cant.Pres.":  _fmt_cant(r["CantPres"]),
                    "P.U.":        _fmt_gs_plain(r["PU"]),
                    "P.Total":     _fmt_gs_plain(r["PTotal"]),
                    "Cant.Ant.":   _fmt_cant(r["CantAnt"]),
                    "Cant.Actual": r["CantActual"],
                    "Cant.Acum.":  _fmt_cant(r["CantAcum"]),
                    "M.Ant.":      _fmt_gs_plain(r["MontoAnt"]),
                    "M.Actual":    _fmt_gs_plain(r["MontoActual"]),
                    "M.Acum.":     _fmt_gs_plain(r["MontoAcum"]),
                    "% Acum.":     r["PctAcum"],
                }
                for r in g_rows
            ])

            if locked:
                st.dataframe(df_section, use_container_width=True, hide_index=True, column_config=_COL_CFG)
            else:
                edited_section = st.data_editor(
                    df_section, use_container_width=True, hide_index=True,
                    disabled=_DISABLED,
                    column_config=_COL_CFG,
                    key=f"cert_editor_{editor_idx}",
                )
                for i, row in edited_section.iterrows():
                    pl_id = g_rows[int(i)]["id"]
                    cant  = float(row.get("Cant.Actual") or 0)
                    if cant:
                        new_edit[pl_id] = cant
            editor_idx += 1

    if not locked:
        if new_edit != st.session_state["cert_lineas_edit"]:
            st.session_state["cert_lineas_edit"] = new_edit
            st.session_state["cert_dirty"] = True
            display_rows = _build_display_rows(presup_lineas, new_edit, anteriores)

    # Overcertification warnings
    warnings = _check_overcertification(display_rows)
    if warnings:
        with st.expander(f"⚠ {len(warnings)} ítem(s) sobre-certificados", expanded=True):
            for w in warnings:
                st.warning(w)

    # Summary metrics
    total_pres    = sum(r["PTotal"] for r in display_rows)
    total_m_act   = sum(r["MontoActual"] for r in display_rows)
    total_m_acu   = sum(r["MontoAcum"] for r in display_rows)
    pct_gral      = (total_m_acu / total_pres * 100) if total_pres else 0

    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("Cert. actual", _fmt_gs(total_m_act))
    mc2.metric("Acumulado", _fmt_gs(total_m_acu))
    mc3.metric("% Avance general", _fmt_pct(pct_gral))
    with mc4:
        if st.session_state["cert_dirty"]:
            st.markdown('<span class="cert-dirty">● Cambios sin guardar</span>', unsafe_allow_html=True)
        elif st.session_state["cert_saved_at"]:
            ts = st.session_state["cert_saved_at"].strftime("%H:%M:%S")
            st.markdown(f'<span class="cert-saved">Guardado a las {ts}</span>', unsafe_allow_html=True)

    # Action buttons
    st.divider()
    if locked:
        bc1, bc2 = st.columns([1, 5])
        with bc1:
            if st.button("🔓 Reabrir", use_container_width=True):
                if _reabrir_cert():
                    st.session_state["cert_pending_delete"] = False
                    st.success("Certificación reabierta.")
                    st.rerun()
    else:
        bc1, bc2, bc3, bc4 = st.columns([1, 1, 3, 1])
        with bc1:
            if st.button("💾 Guardar borrador", use_container_width=True, type="primary"):
                if _guardar_cert(obra_id, presup_lineas):
                    st.success("Borrador guardado.")
                    st.rerun()
        with bc2:
            errors = _validate_cert(cab, st.session_state["cert_lineas_edit"])
            if errors:
                st.button("✅ Confirmar", use_container_width=True, disabled=True)
                for e in errors:
                    st.error(e)
            else:
                if st.button("✅ Confirmar", use_container_width=True):
                    if _confirmar_cert(obra_id, presup_lineas):
                        st.success("Certificación confirmada.")
                        st.rerun()
        with bc4:
            if st.button("🗑️ Eliminar", use_container_width=True):
                st.session_state["cert_pending_delete"] = True
                st.rerun()

    # ── Confirmation zone (only Borrador) ────────────────────────────────────
    if not locked and st.session_state.get("cert_pending_delete"):
        n_lineas = len([
            cl for cl in cert_lineas_all
            if cl.get("cabecera_id") == st.session_state["cert_cab_id"]
        ])
        st.divider()
        st.error(
            f"**Estás por eliminar la Certificación Nº {cert_numero}** de forma permanente.\n\n"
            f"Esto borrará **{n_lineas} línea(s) de certificación** y la cabecera. "
            f"Esta acción **no se puede deshacer**."
        )
        confirmado = st.checkbox(
            "Entiendo que esta acción es irreversible y quiero continuar.",
            key="cert_delete_confirm_chk",
        )
        dc1, dc2, _ = st.columns([1, 1, 4])
        with dc1:
            if st.button(
                "Eliminar definitivamente",
                type="primary",
                disabled=not confirmado,
                use_container_width=True,
            ):
                if _eliminar_cert(cert_lineas_all):
                    st.success("Certificación eliminada correctamente.")
                    st.rerun()
        with dc2:
            if st.button("Cancelar", use_container_width=True):
                st.session_state["cert_pending_delete"] = False
                st.rerun()

    # --- Export section ---
    st.divider()
    st.markdown("##### Exportar")

    obra_rec   = next((o for o in obras_raw if o["id"] == obra_id), {})
    obra_nombre = obra_rec.get("nombre", obra_rec.get("clave", ""))
    fecha_cert  = ""
    fecha_val   = cab.get("fecha_certificado")
    if isinstance(fecha_val, date):
        fecha_cert = fecha_val.strftime("%d/%m/%Y")
    elif isinstance(fecha_val, str) and fecha_val:
        try:
            fecha_cert = datetime.strptime(fecha_val[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            fecha_cert = fecha_val

    col_pdf_memo, col_pdf_det, col_xls = st.columns(3)

    with col_pdf_memo:
        with st.expander("📄 Certificado Memo"):
            prof_name = st.text_input(
                "De (profesional)", value="Arq. José María Sanchez",
                key="cert_memo_de",
            )
            dest_text = st.text_input(
                "A (destinatarios)", value="",
                key="cert_memo_a",
            )
            obras_con_presup = [
                o for o in obras_raw
                if _lineas_for_obra(o["id"], all_presup)
            ]
            multi_opts = {
                o.get("clave", o.get("nombre", "")): o["id"]
                for o in obras_con_presup
            }
            sel_obras_memo = st.multiselect(
                "Obras a incluir",
                list(multi_opts.keys()),
                default=[obra_rec.get("clave", obra_rec.get("nombre", ""))],
                key="cert_memo_obras",
            )
            try:
                obras_data = []
                for obra_label in sel_obras_memo:
                    o_id = multi_opts[obra_label]
                    o_presup = _lineas_for_obra(o_id, all_presup)
                    o_cabs = _cabs_for_obra(o_id, cabeceras_raw)
                    o_cert_num = cert_numero if o_id == obra_id else (
                        max((int(c.get("numero") or 0) for c in o_cabs), default=0)
                    )
                    o_ant = _compute_anteriores(o_id, o_cert_num, cabeceras_raw, cert_lineas_all)
                    o_cert_edit = dict(cert_lineas_ed) if o_id == obra_id else {}
                    if o_id != obra_id:
                        for c in o_cabs:
                            if int(c.get("numero") or 0) == o_cert_num:
                                for cl in cert_lineas_all:
                                    if cl.get("cabecera_id") == c["id"]:
                                        pl_id = cl.get("presupuesto_linea_id")
                                        o_cert_edit[pl_id] = float(cl.get("cantidad_certificada") or 0)
                    o_display = _build_display_rows(o_presup, o_cert_edit, o_ant)
                    o_rec = next((o for o in obras_raw if o["id"] == o_id), {})
                    total_pres_o = sum(r["PTotal"] for r in o_display)
                    total_act_o  = sum(r["MontoActual"] for r in o_display)
                    obras_data.append({
                        "obra_nombre": obra_label,
                        "superficie":  o_rec.get("superficie", ""),
                        "pct_avance":  (total_act_o / total_pres_o * 100) if total_pres_o else 0,
                        "items":       [r for r in o_display if r["CantActual"] > 0],
                        "total":       total_act_o,
                    })
                memo_bytes = generate_certificado_memo_pdf(
                    profesional_nombre=prof_name,
                    destinatarios=dest_text,
                    titulo_medicion=f"CERTIFICADO DE OBRA\n({cert_numero}° MEDICIÓN)",
                    fecha=fecha_cert,
                    obras_data=obras_data,
                    fecha_generacion=date.today().strftime("%d/%m/%Y"),
                )
                st.download_button(
                    "📄 Descargar Memo PDF",
                    data=memo_bytes,
                    file_name=f"Certificado Obra - {obra_nombre} - Cert {cert_numero}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="dl_memo",
                )
            except Exception as e:
                st.error(f"Error al generar PDF memo: {e}")

    with col_pdf_det:
        try:
            cliente_id     = obra_rec.get("cliente_id")
            cliente_nombre = maps["clientes"].get(cliente_id, "")
            det_bytes = generate_certificado_detalle_pdf(
                obra_nombre=obra_nombre,
                ubicacion=obra_rec.get("ubicacion", ""),
                propietario=cliente_nombre,
                profesional="Arq. José María Sánchez G.",
                superficie=obra_rec.get("superficie", ""),
                cert_numero=cert_numero,
                fecha_presupuesto="",
                fecha_certificado=fecha_cert,
                display_rows=display_rows,
                fecha_generacion=date.today().strftime("%d/%m/%Y"),
            )
            st.download_button(
                "📊 Planilla Detallada PDF",
                data=det_bytes,
                file_name=f"Planilla Detallada - {obra_nombre} - Cert {cert_numero}.pdf",
                mime="application/pdf",
                use_container_width=True,
                key="dl_det",
            )
        except Exception as e:
            st.error(f"Error al generar PDF detallada: {e}")

    with col_xls:
        try:
            xlsx_bytes = _generate_cert_excel(obra_nombre, cert_numero, fecha_cert, display_rows)
            st.download_button(
                "📗 Descargar Excel",
                data=xlsx_bytes,
                file_name=f"Certificado - {obra_nombre} - Cert {cert_numero}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Error al generar Excel: {e}")


# ---------------------------------------------------------------------------
# Tab 3 — Resumen financiero por obra
# ---------------------------------------------------------------------------

def _balance_badge(monto: float, label_pos: str, label_neg: str, label_zero: str = "Equilibrado") -> str:
    if monto > 0:
        color, label = "#E67E22", label_pos
    elif monto < 0:
        color, label = "#E74C3C", label_neg
    else:
        color, label = "#27AE60", label_zero
    return (
        f'<span style="color:{color};font-size:.82rem;font-weight:600">'
        f'{label}</span>'
    )


def _render_tab_resumen(obras_raw: list[dict], maps: dict):
    st.markdown(
        '<p class="js-sub">Cuadre financiero por obra: presupuesto contratado, '
        'certificaciones confirmadas y cobros recibidos.</p>',
        unsafe_allow_html=True,
    )

    all_presup      = get_all_records("op_cert_presupuesto_linea")
    cabeceras_raw   = get_all_records("op_cert_cabecera")
    cert_lineas_all = get_all_records("op_cert_linea")
    all_ingresos    = get_all_records("fact_ingreso")

    # Only obras that have at least one presupuesto line
    obras_ids_con_presup = {l.get("obra_id") for l in all_presup if l.get("obra_id")}
    obras_con_presup = [o for o in obras_raw if o["id"] in obras_ids_con_presup]

    if not obras_con_presup:
        st.info("No hay obras con presupuesto definido. Creá uno en la pestaña Presupuesto.")
        return

    obra_id = _render_obra_selector("res", obras_con_presup)
    if not obra_id:
        return

    presup_lineas = _lineas_for_obra(obra_id, all_presup)
    obra_cabs     = _cabs_for_obra(obra_id, cabeceras_raw)
    ingresos      = _ingresos_for_obra(obra_id, all_ingresos)

    # ── Totals ──────────────────────────────────────────────────────────────
    total_pres = sum(
        float(l.get("cantidad") or 0) * float(l.get("precio_unitario") or 0)
        for l in presup_lineas
    )

    # Cert amounts: only Confirmado certs
    pl_pu_map: dict = {
        l["id"]: float(l.get("precio_unitario") or 0) for l in presup_lineas
    }
    confirmed_cab_ids = {c["id"] for c in obra_cabs if c.get("estado") == "Confirmado"}
    cert_amounts: dict = {}
    for cl in cert_lineas_all:
        cab_id = cl.get("cabecera_id")
        if cab_id not in confirmed_cab_ids:
            continue
        pl_id = cl.get("presupuesto_linea_id")
        cant  = float(cl.get("cantidad_certificada") or 0)
        cert_amounts[cab_id] = cert_amounts.get(cab_id, 0) + cant * pl_pu_map.get(pl_id, 0)

    total_cert    = sum(cert_amounts.values())
    total_cobrado = sum(float(i.get("monto_recibido") or 0) for i in ingresos)

    saldo_pendiente = total_pres - total_cobrado   # lo que falta cobrar del contrato
    balance         = total_cobrado - total_cert   # cobrado vs certificado

    pct_cert = (total_cert / total_pres * 100) if total_pres else 0

    # ── Metric cards ────────────────────────────────────────────────────────
    mc1, mc2, mc3 = st.columns(3)
    mc1.metric("Presupuesto contratado", _fmt_gs(total_pres))
    mc2.metric("Total certificado", _fmt_gs(total_cert), delta=f"{_fmt_pct(pct_cert)} del contrato")
    mc3.metric("Total cobrado", _fmt_gs(total_cobrado))

    st.divider()

    # ── Balance indicators ──────────────────────────────────────────────────
    bc1, bc2 = st.columns(2)
    with bc1:
        st.metric("Saldo pendiente (Ppto − Cobrado)", _fmt_gs(abs(saldo_pendiente)))
        st.markdown(
            _balance_badge(
                saldo_pendiente,
                label_pos="Por cobrar del cliente",
                label_neg="Cobrado en exceso del contrato",
                label_zero="Contrato saldado",
            ),
            unsafe_allow_html=True,
        )
    with bc2:
        st.metric("Balance (Cobrado − Certificado)", _fmt_gs(abs(balance)))
        st.markdown(
            _balance_badge(
                balance,
                label_pos="Cobrado por delante de lo certificado",
                label_neg="Certificado por delante de los cobros",
                label_zero="Cobrado = Certificado",
            ),
            unsafe_allow_html=True,
        )

    # ── Timeline ────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("##### Movimientos")

    # Presupuesto row (fija al inicio, sin fecha de evento)
    timeline: list[dict] = [
        {
            "_sort":        "",
            "Fecha":        "",
            "Tipo":         "Presupuesto",
            "Concepto":     "Presupuesto contratado",
            "Presupuesto":  total_pres,
            "Cobrado":      None,
            "Certificado":  None,
        }
    ]

    for c in obra_cabs:
        num    = c.get("numero", "?")
        iso    = str(c.get("fecha_certificado") or "")[:10]
        estado = c.get("estado", "")
        monto  = cert_amounts.get(c["id"], 0)
        obs    = c.get("observaciones", "") or ""
        concepto = f"Cert. Nº {num}"
        if obs:
            concepto += f" — {obs}"
        timeline.append({
            "_sort":       iso,
            "Fecha":       _fmt_fecha(iso),
            "Tipo":        "Certificación" + (" [Borrador]" if estado == "Borrador" else ""),
            "Concepto":    concepto,
            "Presupuesto": None,
            "Cobrado":     None,
            "Certificado": monto if monto else None,
        })

    for ing in ingresos:
        iso      = str(ing.get("fecha_ingreso") or "")[:10]
        tipo     = ing.get("tipo_ingreso") or "Ingreso"
        concepto = ing.get("concepto") or ""
        monto    = float(ing.get("monto_recibido") or 0)
        timeline.append({
            "_sort":       iso,
            "Fecha":       _fmt_fecha(iso),
            "Tipo":        tipo,
            "Concepto":    concepto,
            "Presupuesto": None,
            "Cobrado":     monto if monto else None,
            "Certificado": None,
        })

    # Sort by date, keeping presupuesto row first (_sort == "")
    timeline.sort(key=lambda x: x["_sort"] or "")

    def _tl_gs(v):
        return _fmt_gs_plain(v) if v is not None else ""

    df_tl = pd.DataFrame([
        {
            "Fecha":        r["Fecha"],
            "Tipo":         r["Tipo"],
            "Concepto":     r["Concepto"],
            "Presupuesto":  _tl_gs(r.get("Presupuesto")),
            "Cobrado":      _tl_gs(r.get("Cobrado")),
            "Certificado":  _tl_gs(r.get("Certificado")),
        }
        for r in timeline
    ])

    st.dataframe(
        df_tl, use_container_width=True, hide_index=True,
        column_config={
            "Fecha":        st.column_config.TextColumn(width="small"),
            "Tipo":         st.column_config.TextColumn(width="medium"),
            "Concepto":     st.column_config.TextColumn(width="large"),
            "Presupuesto":  st.column_config.TextColumn("Presupuesto (Gs.)", width="medium"),
            "Cobrado":      st.column_config.TextColumn("Cobrado (Gs.)", width="medium"),
            "Certificado":  st.column_config.TextColumn("Certificado (Gs.)", width="medium"),
        },
    )

    # Totals at bottom
    total_cert_shown    = sum(r.get("Certificado") or 0 for r in timeline)
    total_cobrado_shown = sum(r.get("Cobrado") or 0 for r in timeline)
    st.markdown(
        f'<div style="text-align:right;font-size:.82rem;color:#6B7280;margin-top:4px">'
        f'Total certificado: <b>{_fmt_gs(total_cert_shown)}</b>'
        f'&nbsp;&nbsp;|&nbsp;&nbsp;'
        f'Total cobrado: <b>{_fmt_gs(total_cobrado_shown)}</b>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ── Export ──────────────────────────────────────────────────────────────
    st.divider()
    obra_rec    = next((o for o in obras_raw if o["id"] == obra_id), {})
    obra_nombre = obra_rec.get("clave", obra_rec.get("nombre", ""))
    cliente_id  = obra_rec.get("cliente_id")
    cliente_str = maps["clientes"].get(cliente_id, "")

    try:
        resumen_pdf_bytes = generate_resumen_financiero_pdf(
            obra_nombre=obra_nombre,
            cliente=cliente_str,
            fecha_generacion=date.today().strftime("%d/%m/%y"),
            total_presupuesto=total_pres,
            total_certificado=total_cert,
            total_cobrado=total_cobrado,
            saldo_pendiente=saldo_pendiente,
            balance=balance,
            timeline=timeline,
        )
        st.download_button(
            "📄 Exportar PDF",
            data=resumen_pdf_bytes,
            file_name=f"Resumen Financiero - {obra_nombre}.pdf",
            mime="application/pdf",
        )
    except Exception as e:
        st.error(f"Error al generar PDF: {e}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    st.set_page_config(page_title="Certificaciones — JS Tools", layout="wide")
    st.markdown(_CSS, unsafe_allow_html=True)
    _init_session_state()

    st.title("📋 Certificaciones y Avances de Obra")

    obras_raw = get_all_records("dim_obra")
    cli_recs  = get_all_records("dim_cliente")
    maps = {
        "clientes": {r["id"]: r.get("nombre_cliente", str(r["id"])) for r in cli_recs},
    }

    tab1, tab2, tab3 = st.tabs(["📝 Presupuesto", "📋 Certificaciones", "📊 Resumen"])

    with tab1:
        _render_tab_presupuesto(obras_raw)

    with tab2:
        _render_tab_certificaciones(obras_raw, maps)

    with tab3:
        _render_tab_resumen(obras_raw, maps)


main()
