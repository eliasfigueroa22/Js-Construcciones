"""Tool 5 — Planilla de Mediciones.

Permite a los arquitectos cargar mediciones semanales de obra por subcontratista,
calcular totales automáticamente y exportar la planilla a PDF/Excel.
Los datos se sincronizan con Supabase (tablas op_medicion_cabecera + op_medicion_linea).
"""

from __future__ import annotations

import io
from datetime import date, datetime

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from connectors.supabase_connector import (
    clear_cache,
    create_record,
    delete_record,
    get_all_records,
    update_record,
)
from core.base_tool import ToolMetadata
from generators.pdf_generator import generate_acumulado_pdf, generate_medicion_pdf

TOOL = ToolMetadata(
    name="Mediciones",
    description="Planilla de cómputo métrico semanal por obra y subcontratista.",
    icon="📐",
    page_file="05_mediciones.py",
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_UNIDAD_OPTIONS = ["m²", "m³", "ml", "kg", "un"]

# ---------------------------------------------------------------------------
# CSS
# ---------------------------------------------------------------------------

_CSS = """
<style>
:root {
    --js-accent:  #E8622A;
    --js-success: #27AE60;
    --js-danger:  #E74C3C;
    --js-warn:    #E67E22;
    --js-muted:   #6B7280;
    --js-border:  rgba(255,255,255,0.09);
    --js-surface: rgba(255,255,255,0.03);
}
.js-sub  { color: var(--js-muted); font-size: .875rem; margin-top: -10px; margin-bottom: 24px; }
.js-pill { display: inline-block; border-radius: 3px; padding: 2px 8px; font-size: .68rem;
           font-weight: 700; letter-spacing: .8px; text-transform: uppercase; }
.med-pill-borrador  { background: rgba(230,126,34,.18); color: #E67E22; }
.med-pill-confirmado{ background: rgba(39,174,96,.18);  color: #27AE60; }
.med-dirty  { color: var(--js-warn); font-size: .82rem; font-weight: 600; }
.med-saved  { color: var(--js-muted); font-size: .82rem; }
.med-sin-precio { background: rgba(231,76,60,.12); color: #E74C3C;
                  border-radius: 4px; padding: 4px 10px; font-size: .82rem; }
</style>
"""

# ---------------------------------------------------------------------------
# Lookup helpers
# ---------------------------------------------------------------------------

def _sectores_for_obra(obra_id: int, sectores_raw: list[dict]) -> list[dict]:
    """Filtra sectores que pertenecen a la obra dada."""
    if not obra_id:
        return []
    return [s for s in sectores_raw if s.get("obra_id") == obra_id]


def _id_to_name(records: list[dict], name_key: str) -> dict[int, str]:
    return {r["id"]: str(r.get(name_key, "")) for r in records}


def _name_to_id(records: list[dict], name_key: str) -> dict[str, int]:
    return {str(r.get(name_key, "")): r["id"] for r in records}


# ---------------------------------------------------------------------------
# Session state
# ---------------------------------------------------------------------------

_SS_KEYS = {
    "med_cabecera_id":  None,
    "med_cabecera":     None,
    "med_lineas":       None,
    "med_deleted_ids":  None,
    "med_dirty":        False,
    "med_saved_at":     None,
}

def _init_session_state():
    for key, default in _SS_KEYS.items():
        if key not in st.session_state:
            st.session_state[key] = [] if default is None and key in ("med_lineas", "med_deleted_ids") else default
    if st.session_state["med_cabecera"] is None:
        st.session_state["med_cabecera"] = {
            "obra_id": None, "trabajador_id": None,
            "fecha": None, "estado": "Borrador", "observaciones": "",
        }
    if st.session_state["med_lineas"] is None:
        st.session_state["med_lineas"] = []
    if st.session_state["med_deleted_ids"] is None:
        st.session_state["med_deleted_ids"] = []


def _reset_session_state():
    st.session_state["med_cabecera_id"]  = None
    st.session_state["med_cabecera"]     = {
        "obra_id": None, "trabajador_id": None,
        "fecha": None, "estado": "Borrador", "observaciones": "",
    }
    st.session_state["med_lineas"]       = []
    st.session_state["med_deleted_ids"]  = []
    st.session_state["med_dirty"]        = False
    st.session_state["med_saved_at"]     = None


def _mark_dirty():
    st.session_state["med_dirty"] = True


def _load_borrador_into_state(cabecera: dict, lineas: list[dict]):
    st.session_state["med_cabecera_id"] = cabecera["id"]
    st.session_state["med_cabecera"] = {
        "obra_id":        cabecera.get("obra_id"),
        "trabajador_id":  cabecera.get("trabajador_id"),
        "fecha":          cabecera.get("fecha"),
        "estado":         cabecera.get("estado", "Borrador"),
        "observaciones":  cabecera.get("observaciones", ""),
    }
    st.session_state["med_lineas"] = [
        {
            "id":             l["id"],
            "sector_id":      l.get("sector_id"),
            "rubro_id":       l.get("rubro_id"),
            "descripcion":    l.get("descripcion", ""),
            "unidad":         l.get("unidad", ""),
            "largo":          float(l.get("largo") or 0),
            "ancho":          float(l.get("ancho") or 0),
            "alto":           float(l.get("alto") or 0),
            "cantidad":       float(l.get("cantidad") or 0),
            "precio_unitario": float(l.get("precio_unitario") or 0),
        }
        for l in lineas
    ]
    st.session_state["med_deleted_ids"] = []
    st.session_state["med_dirty"]       = False
    st.session_state["med_saved_at"]    = None


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def _validate_cabecera(cab: dict) -> list[str]:
    errors = []
    if not cab.get("obra_id"):
        errors.append("Seleccioná una obra.")
    if not cab.get("trabajador_id"):
        errors.append("Seleccioná un subcontratista.")
    if not cab.get("fecha"):
        errors.append("Ingresá la fecha de medición.")
    return errors


def _validate_lineas(lineas: list[dict]) -> tuple[list[str], int]:
    errors = []
    n_sin_precio = 0
    if not lineas:
        errors.append("Agregá al menos una línea de medición.")
        return errors, 0
    for i, l in enumerate(lineas, 1):
        if not l.get("sector_id"):
            errors.append(f"Línea {i}: falta sector.")
        if not l.get("rubro_id"):
            errors.append(f"Línea {i}: falta rubro.")
        if not str(l.get("descripcion", "")).strip():
            errors.append(f"Línea {i}: falta descripción.")
        if float(l.get("cantidad") or 0) <= 0:
            errors.append(f"Línea {i}: la cantidad debe ser mayor a 0.")
        if float(l.get("precio_unitario") or 0) == 0:
            n_sin_precio += 1
    return errors, n_sin_precio


def _check_duplicate_confirmada(
    obra_id: int | None, trab_id: int | None, fecha: date | str | None,
    cabeceras_raw: list[dict], current_id: int | None,
) -> bool:
    """True si ya existe una Confirmada para la misma obra+subcontratista en la misma semana ISO."""
    if not fecha or not obra_id or not trab_id:
        return False
    try:
        if isinstance(fecha, str):
            fecha = datetime.strptime(fecha[:10], "%Y-%m-%d").date()
        week = fecha.isocalendar()[:2]  # (year, week)
    except (ValueError, AttributeError):
        return False

    for cab in cabeceras_raw:
        if cab["id"] == current_id:
            continue
        if cab.get("estado") != "Confirmado":
            continue
        if cab.get("obra_id") != obra_id:
            continue
        if cab.get("trabajador_id") != trab_id:
            continue
        try:
            cab_fecha = cab.get("fecha", "")
            if isinstance(cab_fecha, str):
                cab_fecha = datetime.strptime(cab_fecha[:10], "%Y-%m-%d").date()
            if cab_fecha.isocalendar()[:2] == week:
                return True
        except (ValueError, AttributeError):
            continue
    return False


# ---------------------------------------------------------------------------
# DataFrame <-> session state conversion
# ---------------------------------------------------------------------------

def _lineas_to_df(
    lineas: list[dict],
    sec_id_name: dict[int, str],
    rub_id_name: dict[int, str],
) -> pd.DataFrame:
    """Convierte med_lineas (IDs) a DataFrame de display (nombres)."""
    rows = []
    for l in lineas:
        cant = float(l.get("cantidad") or 0)
        pu   = float(l.get("precio_unitario") or 0)
        rows.append({
            "Sector":          sec_id_name.get(l.get("sector_id"), ""),
            "Rubro":           rub_id_name.get(l.get("rubro_id"), ""),
            "Descripcion":     l.get("descripcion", ""),
            "Unidad":          l.get("unidad", ""),
            "Largo":           float(l.get("largo") or 0),
            "Ancho":           float(l.get("ancho") or 0),
            "Alto":            float(l.get("alto") or 0),
            "Cantidad":        cant,
            "PrecioUnitario":  pu,
            "Total":           cant * pu,
        })
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        "Sector", "Rubro", "Descripcion", "Unidad",
        "Largo", "Ancho", "Alto", "Cantidad", "PrecioUnitario", "Total",
    ])


def _df_to_lineas(
    df: pd.DataFrame,
    prev_lineas: list[dict],
    sec_name_id: dict[str, int],
    rub_name_id: dict[str, int],
    deleted_ids: list[int],
) -> tuple[list[dict], list[int]]:
    """
    Convierte DataFrame editado de vuelta a lista de dicts con IDs.
    Detecta filas eliminadas y agrega sus id a deleted_ids.
    Retorna (nuevas_lineas, nuevos_deleted_ids).
    """
    new_deleted = list(deleted_ids)

    # Detect deleted rows: prev_lineas has more rows than df
    if len(prev_lineas) > len(df):
        for i in range(len(df), len(prev_lineas)):
            rec_id = prev_lineas[i].get("id")
            if rec_id and rec_id not in new_deleted:
                new_deleted.append(rec_id)

    new_lineas = []
    for i, row in df.iterrows():
        idx = int(i)
        prev = prev_lineas[idx] if idx < len(prev_lineas) else {}

        sector_name = str(row.get("Sector", "") or "")
        rubro_name  = str(row.get("Rubro", "") or "")

        # Sector: try name→id; fallback to prev id if name not found (stale after obra change)
        sector_id = sec_name_id.get(sector_name) or prev.get("sector_id")
        rubro_id  = rub_name_id.get(rubro_name)  or prev.get("rubro_id")

        largo = float(row.get("Largo") or 0)
        ancho = float(row.get("Ancho") or 0)
        alto  = float(row.get("Alto")  or 0)

        # Auto-calc Cantidad if any dimension is provided
        if largo > 0 or ancho > 0 or alto:
            l = largo if largo > 0 else 1
            a = ancho if ancho > 0 else 1
            h = alto  if alto  > 0 else 1
            cantidad = l * a * h
        else:
            cantidad = float(row.get("Cantidad") or 0)

        new_lineas.append({
            "id":              prev.get("id"),
            "sector_id":       sector_id,
            "rubro_id":        rubro_id,
            "descripcion":     str(row.get("Descripcion", "") or "").strip(),
            "unidad":          str(row.get("Unidad", "") or ""),
            "largo":           largo,
            "ancho":           ancho,
            "alto":            alto,
            "cantidad":        round(cantidad, 3),
            "precio_unitario": float(row.get("PrecioUnitario") or 0),
        })

    return new_lineas, new_deleted


# ---------------------------------------------------------------------------
# Supabase write operations
# ---------------------------------------------------------------------------

def _guardar(cabeceras_raw: list[dict]) -> bool:
    """Guarda cabecera + líneas en Supabase. Retorna True si exitoso."""
    cab    = st.session_state["med_cabecera"]
    lineas = st.session_state["med_lineas"]

    # Fase 1: eliminar líneas borradas
    for rec_id in st.session_state["med_deleted_ids"]:
        try:
            delete_record("op_medicion_linea", rec_id)
        except ConnectionError as e:
            st.error(f"Error al eliminar línea: {e}")
            return False

    # Fase 2: upsert cabecera
    fecha_val = cab.get("fecha")
    if isinstance(fecha_val, date):
        fecha_str = fecha_val.isoformat()
    elif isinstance(fecha_val, str):
        fecha_str = fecha_val[:10]
    else:
        fecha_str = ""

    cab_fields = {
        "obra_id":        cab.get("obra_id"),
        "trabajador_id":  cab.get("trabajador_id"),
        "fecha":          fecha_str,
        "estado":         cab.get("estado", "Borrador"),
        "observaciones":  cab.get("observaciones", ""),
    }

    try:
        if st.session_state["med_cabecera_id"]:
            update_record("op_medicion_cabecera", st.session_state["med_cabecera_id"], cab_fields)
        else:
            new_id = create_record("op_medicion_cabecera", cab_fields)
            st.session_state["med_cabecera_id"] = new_id
    except ConnectionError as e:
        st.error(f"Error al guardar cabecera: {e}")
        return False

    cab_id = st.session_state["med_cabecera_id"]

    # Fase 3: upsert líneas
    for linea in lineas:
        linea_fields = {
            "cabecera_id":    cab_id,
            "sector_id":      linea.get("sector_id"),
            "rubro_id":       linea.get("rubro_id"),
            "descripcion":    linea.get("descripcion", ""),
            "unidad":         linea.get("unidad", ""),
            "largo":          linea.get("largo") or None,
            "ancho":          linea.get("ancho") or None,
            "alto":           linea.get("alto")  or None,
            "cantidad":       linea.get("cantidad") or None,
            "precio_unitario": linea.get("precio_unitario") or None,
        }
        try:
            if linea.get("id"):
                update_record("op_medicion_linea", linea["id"], linea_fields)
            else:
                new_id = create_record("op_medicion_linea", linea_fields)
                linea["id"] = new_id
        except ConnectionError as e:
            st.error(f"Error al guardar línea: {e}")
            return False

    # Éxito
    st.session_state["med_deleted_ids"] = []
    st.session_state["med_dirty"]       = False
    st.session_state["med_saved_at"]    = datetime.now()
    clear_cache()
    return True


def _confirmar(cabeceras_raw: list[dict]) -> bool:
    if not _guardar(cabeceras_raw):
        return False
    try:
        update_record("op_medicion_cabecera", st.session_state["med_cabecera_id"], {"estado": "Confirmado"})
    except ConnectionError as e:
        st.error(f"Error al confirmar: {e}")
        return False
    st.session_state["med_cabecera"]["estado"] = "Confirmado"
    st.session_state["med_dirty"] = False
    clear_cache()
    return True


def _reabrir() -> bool:
    try:
        update_record("op_medicion_cabecera", st.session_state["med_cabecera_id"], {"estado": "Borrador"})
    except ConnectionError as e:
        st.error(f"Error al reabrir: {e}")
        return False
    st.session_state["med_cabecera"]["estado"] = "Borrador"
    clear_cache()
    return True


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _generate_excel(
    cab: dict,
    lineas: list[dict],
    sec_id_name: dict[int, str],
    rub_id_name: dict[int, str],
    obra_nombre: str,
    trab_nombre: str,
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Medición"

    accent = "E8622A"
    dark    = "1A252F"
    gray    = "F2F3F4"

    hdr_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hdr_fill   = PatternFill("solid", fgColor=accent)
    title_font = Font(name="Calibri", bold=True, color=dark, size=12)
    sub_font   = Font(name="Calibri", color="5D6D7E", size=9)
    bold_font  = Font(name="Calibri", bold=True, size=10)
    normal_font= Font(name="Calibri", size=10)
    total_fill = PatternFill("solid", fgColor="AAB7B8")
    alt_fill   = PatternFill("solid", fgColor="F2F3F4")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fecha_val = cab.get("fecha", "")
    if isinstance(fecha_val, date):
        fecha_str = fecha_val.strftime("%d/%m/%Y")
    else:
        fecha_str = str(fecha_val)[:10] if fecha_val else ""

    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = "PLANILLA DE MEDICIÓN"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Obra: {obra_nombre}  |  Subcontratista: {trab_nombre}  |  Fecha: {fecha_str}"
    ws["A2"].font = sub_font
    ws["A2"].alignment = Alignment(horizontal="center")

    if cab.get("observaciones"):
        ws.merge_cells("A3:J3")
        ws["A3"] = f"Observaciones: {cab['observaciones']}"
        ws["A3"].font = sub_font
        ws["A3"].alignment = Alignment(horizontal="left")
        hdr_row = 5
    else:
        hdr_row = 4

    # Header
    headers = ["Sector", "Rubro", "Descripción", "Unidad", "Largo", "Ancho", "Alto", "Cantidad", "P.U. (Gs.)", "Total (Gs.)"]
    col_widths = [18, 10, 35, 8, 8, 8, 8, 10, 14, 14]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=hdr_row, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[hdr_row].height = 18
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    # Data rows
    grand_total = 0.0
    for i, l in enumerate(lineas):
        row_num = hdr_row + 1 + i
        cant = float(l.get("cantidad") or 0)
        pu   = float(l.get("precio_unitario") or 0)
        total = cant * pu
        grand_total += total

        sec_name = sec_id_name.get(l.get("sector_id"), "")
        rub_name = rub_id_name.get(l.get("rubro_id"), "")

        values = [
            sec_name,
            rub_name,
            l.get("descripcion", ""),
            l.get("unidad", ""),
            l.get("largo") or None,
            l.get("ancho") or None,
            l.get("alto")  or None,
            cant if cant else None,
            pu if pu else None,
            total if pu else None,
        ]
        fill = alt_fill if i % 2 == 1 else None
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            if ci in (5, 6, 7, 8):
                cell.number_format = "#,##0.000"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif ci in (9, 10):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Grand total row
    total_row = hdr_row + 1 + len(lineas)
    ws.merge_cells(f"A{total_row}:I{total_row}")
    cell_label = ws.cell(row=total_row, column=1, value="TOTAL GENERAL")
    cell_label.font = bold_font
    cell_label.fill = total_fill
    cell_label.alignment = Alignment(horizontal="right", vertical="center")
    cell_label.border = border

    cell_total = ws.cell(row=total_row, column=10, value=grand_total)
    cell_total.font = bold_font
    cell_total.fill = total_fill
    cell_total.number_format = "#,##0"
    cell_total.alignment = Alignment(horizontal="right", vertical="center")
    cell_total.border = border

    ws.auto_filter.ref = f"A{hdr_row}:J{hdr_row + len(lineas)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Aggregation (Vista General)
# ---------------------------------------------------------------------------

def _aggregate_lineas(
    cabeceras: list[dict],
    all_lineas: list[dict],
    rub_id_name: dict[int, str],
) -> tuple[list[dict], int, float]:
    """Agrupa líneas por (rubro_id, descripcion, unidad) y suma cantidades/totales.

    Returns:
        (items_agregados, n_planillas, total_gs_acumulado)
    """
    from collections import defaultdict

    cab_ids = {c["id"] for c in cabeceras}

    # Map cabecera_id → fecha string for sorting (to find "last" price)
    cab_fecha: dict[int, str] = {}
    for c in cabeceras:
        f = c.get("fecha", "")
        cab_fecha[c["id"]] = str(f)[:10] if f else ""

    relevant = [l for l in all_lineas if l.get("cabecera_id") in cab_ids]
    # Sort ascending by fecha so the last entry wins for PUUltimo
    relevant.sort(key=lambda l: cab_fecha.get(l.get("cabecera_id"), ""))

    groups: dict = defaultdict(lambda: {"cant": 0.0, "total": 0.0, "last_pu": 0.0, "n": 0})

    for l in relevant:
        rubro_id = l.get("rubro_id")
        rubro_nm = rub_id_name.get(rubro_id, "—")
        key = (rubro_id, rubro_nm, (l.get("descripcion") or "").strip(), l.get("unidad") or "")

        cant = float(l.get("cantidad") or 0)
        pu   = float(l.get("precio_unitario") or 0)
        groups[key]["cant"]  += cant
        groups[key]["total"] += cant * pu
        groups[key]["n"]     += 1
        if pu:
            groups[key]["last_pu"] = pu

    items = []
    total_gs = 0.0
    for (rubro_id, rubro_nm, desc, unidad), data in sorted(groups.items(), key=lambda x: (x[0][1], x[0][2])):
        items.append({
            "_rubro_id":      rubro_id,
            "_rubro_name":    rubro_nm,
            "Descripcion":    desc,
            "Unidad":         unidad,
            "CantidadTotal":  round(data["cant"], 3),
            "PUUltimo":       data["last_pu"],
            "TotalAcumulado": data["total"],
            "Mediciones":     data["n"],
        })
        total_gs += data["total"]

    return items, len(cab_ids), total_gs


# ---------------------------------------------------------------------------
# Excel export — acumulado
# ---------------------------------------------------------------------------

def _generate_acumulado_excel(
    items: list[dict],
    obra_nombre: str,
    trab_nombre: str,
    periodo: str,
    n_planillas: int,
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Acumulado"

    accent = "E8622A"
    dark   = "1A252F"

    hdr_font    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hdr_fill    = PatternFill("solid", fgColor=accent)
    title_font  = Font(name="Calibri", bold=True, color=dark, size=12)
    sub_font    = Font(name="Calibri", color="5D6D7E", size=9)
    bold_font   = Font(name="Calibri", bold=True, size=10)
    normal_font = Font(name="Calibri", size=10)
    total_fill  = PatternFill("solid", fgColor="AAB7B8")
    alt_fill    = PatternFill("solid", fgColor="F2F3F4")
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:G1")
    ws["A1"] = "ACUMULADO DE MEDICIONES"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Obra: {obra_nombre}  |  Subcontratista: {trab_nombre}  |  Período: {periodo}"
    ws["A2"].font = sub_font
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:G3")
    ws["A3"] = f"Planillas incluidas: {n_planillas}"
    ws["A3"].font = sub_font
    ws["A3"].alignment = Alignment(horizontal="left")

    hdr_row = 5
    headers    = ["Rubro", "Descripción", "Unidad", "Cant. Total", "P.U. Último (Gs.)", "Total Acum. (Gs.)", "N° Mediciones"]
    col_widths = [12, 45, 8, 12, 20, 20, 14]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=hdr_row, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[hdr_row].height = 18
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    grand_total = 0.0
    for i, item in enumerate(items):
        row_num = hdr_row + 1 + i
        cant  = item.get("CantidadTotal", 0)
        pu    = item.get("PUUltimo") or None
        total = item.get("TotalAcumulado", 0)
        grand_total += float(total)

        values = [
            item.get("_rubro_name", ""),
            item.get("Descripcion", ""),
            item.get("Unidad", ""),
            cant if cant else None,
            int(pu) if pu else None,
            int(total) if total else None,
            item.get("Mediciones", 0),
        ]
        fill = alt_fill if i % 2 == 1 else None
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            if ci == 4:
                cell.number_format = "#,##0.000"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif ci in (5, 6):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    total_row = hdr_row + 1 + len(items)
    ws.merge_cells(f"A{total_row}:E{total_row}")
    lbl = ws.cell(row=total_row, column=1, value="TOTAL ACUMULADO")
    lbl.font = bold_font
    lbl.fill = total_fill
    lbl.alignment = Alignment(horizontal="right", vertical="center")
    lbl.border = border

    tot = ws.cell(row=total_row, column=6, value=int(grand_total))
    tot.font = bold_font
    tot.fill = total_fill
    tot.number_format = "#,##0"
    tot.alignment = Alignment(horizontal="right", vertical="center")
    tot.border = border

    ws.auto_filter.ref = f"A{hdr_row}:G{hdr_row + len(items)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# UI — vista general (tab 2)
# ---------------------------------------------------------------------------

def _render_vista_general(
    obras_raw: list[dict],
    trabajadores_raw: list[dict],
    cabeceras_raw: list[dict],
    all_lineas: list[dict],
    sec_id_name: dict[int, str],
    rub_id_name: dict[int, str],
    obra_id_to_clave: dict[int, str],
    trab_id_to_name: dict[int, str],
):
    st.markdown('<p class="js-sub">Sumariza todas las mediciones según los filtros. Muestra cantidades totales y precio unitario más reciente por ítem.</p>', unsafe_allow_html=True)

    # Filters
    obra_clave_to_id = {v: k for k, v in obra_id_to_clave.items()}
    obra_opts        = ["(Todas las obras)"] + sorted(obra_id_to_clave.values())

    trab_name_to_id = {v: k for k, v in trab_id_to_name.items()}
    trab_opts       = ["(Todos)"] + sorted([r.get("nombre_completo", "") for r in trabajadores_raw if r.get("nombre_completo")])

    c1, c2, c3, c4, c5 = st.columns([2, 2, 1.5, 1.5, 1.2])
    with c1:
        obra_sel = st.selectbox("Obra", obra_opts, key="vg_obra")
    with c2:
        trab_sel = st.selectbox("Subcontratista", trab_opts, key="vg_trab")
    with c3:
        fecha_desde = st.date_input("Desde", value=None, key="vg_desde")
    with c4:
        fecha_hasta = st.date_input("Hasta", value=date.today(), key="vg_hasta")
    with c5:
        solo_conf = st.checkbox("Solo confirmadas", value=True, key="vg_conf")

    # Filter cabeceras
    filtered_cabs = []
    for c in cabeceras_raw:
        if solo_conf and c.get("estado") != "Confirmado":
            continue

        if obra_sel != "(Todas las obras)":
            if c.get("obra_id") != obra_clave_to_id.get(obra_sel):
                continue

        if trab_sel != "(Todos)":
            if c.get("trabajador_id") != trab_name_to_id.get(trab_sel):
                continue

        fecha_c = c.get("fecha", "")
        if isinstance(fecha_c, str):
            try:
                fecha_c = datetime.strptime(fecha_c[:10], "%Y-%m-%d").date()
            except ValueError:
                fecha_c = None
        if fecha_desde and fecha_c and fecha_c < fecha_desde:
            continue
        if fecha_hasta and fecha_c and fecha_c > fecha_hasta:
            continue

        filtered_cabs.append(c)

    if not filtered_cabs:
        st.info("No hay mediciones que coincidan con los filtros.")
        return

    items, n_planillas, total_gs = _aggregate_lineas(filtered_cabs, all_lineas, rub_id_name)

    if not items:
        st.info("No hay líneas de medición en las planillas seleccionadas.")
        return

    # Summary cards
    n_sin_precio = sum(1 for i in items if not i.get("PUUltimo"))
    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("Planillas incluidas", n_planillas)
    mc2.metric("Ítems únicos", len(items))
    mc3.metric("Total acumulado", f"₲ {int(total_gs):,}".replace(",", "."))
    if n_sin_precio:
        mc4.metric("Sin precio definido", n_sin_precio)

    st.divider()

    # Table
    df_items = pd.DataFrame([{
        "Rubro":               i.get("_rubro_name", ""),
        "Descripción":         i.get("Descripcion", ""),
        "Unidad":              i.get("Unidad", ""),
        "Cant. Total":         i.get("CantidadTotal", 0),
        "P.U. Último (Gs.)":   int(i["PUUltimo"]) if i.get("PUUltimo") else None,
        "Total Acum. (Gs.)":   int(i.get("TotalAcumulado") or 0),
        "N° Mediciones":       i.get("Mediciones", 0),
    } for i in items])

    st.dataframe(
        df_items,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Cant. Total":        st.column_config.NumberColumn(format="%.3f"),
            "P.U. Último (Gs.)":  st.column_config.NumberColumn(format="%d"),
            "Total Acum. (Gs.)":  st.column_config.NumberColumn(format="%d"),
        },
    )

    if n_sin_precio:
        st.markdown(
            f'<span class="med-sin-precio">⚠ {n_sin_precio} ítem(s) sin precio definido — '
            f'el total acumulado es parcial.</span>',
            unsafe_allow_html=True,
        )

    # Export
    st.divider()
    st.markdown("##### Exportar acumulado")

    p_desde   = fecha_desde.strftime("%d/%m/%Y") if fecha_desde else "inicio"
    p_hasta   = fecha_hasta.strftime("%d/%m/%Y") if fecha_hasta else date.today().strftime("%d/%m/%Y")
    periodo_str   = f"{p_desde} – {p_hasta}"
    obra_label    = obra_sel if obra_sel != "(Todas las obras)" else "Todas las obras"
    trab_label    = trab_sel if trab_sel != "(Todos)" else "Todos"
    today_str     = date.today().strftime("%d/%m/%Y")
    nombre_archivo = f"Acumulado - {obra_label} - {trab_label}"

    col_pdf, col_xls = st.columns(2)
    with col_pdf:
        try:
            pdf_bytes = generate_acumulado_pdf(
                obra_nombre=obra_label,
                subcontratista_nombre=trab_label,
                periodo=periodo_str,
                n_planillas=n_planillas,
                items=items,
                fecha_generacion=today_str,
            )
            st.download_button(
                "📄 Descargar PDF",
                data=pdf_bytes,
                file_name=f"{nombre_archivo}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Error al generar PDF: {e}")

    with col_xls:
        try:
            xlsx_bytes = _generate_acumulado_excel(items, obra_label, trab_label, periodo_str, n_planillas)
            st.download_button(
                "📊 Descargar Excel",
                data=xlsx_bytes,
                file_name=f"{nombre_archivo}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Error al generar Excel: {e}")


# ---------------------------------------------------------------------------
# UI — load section
# ---------------------------------------------------------------------------

def _render_load_section(
    cabeceras_raw: list[dict],
    obra_id_to_clave: dict[int, str],
    trab_id_to_name: dict[int, str],
    all_lineas_raw: list[dict],
):
    has_active = st.session_state["med_cabecera_id"] is not None
    with st.expander("📂 Cargar medición existente", expanded=not has_active):
        # Filters row
        fc1, fc2, fc3, fc4 = st.columns([2, 2, 1.5, 1])

        # Build obra options from cabeceras that exist
        obra_ids_in_cabs = sorted(
            {obra_id_to_clave.get(c.get("obra_id"), "") for c in cabeceras_raw},
        )
        obra_ids_in_cabs = [o for o in obra_ids_in_cabs if o]

        trab_ids_in_cabs = sorted(
            {trab_id_to_name.get(c.get("trabajador_id"), "") for c in cabeceras_raw},
        )
        trab_ids_in_cabs = [t for t in trab_ids_in_cabs if t]

        with fc1:
            obra_filter = st.selectbox(
                "Obra", ["(Todas)"] + obra_ids_in_cabs, key="med_load_obra",
            )
        with fc2:
            trab_filter = st.selectbox(
                "Subcontratista", ["(Todos)"] + trab_ids_in_cabs, key="med_load_trab",
            )
        with fc3:
            fecha_filter = st.date_input("Desde", value=None, key="med_load_desde")
        with fc4:
            st.markdown("<br>", unsafe_allow_html=True)
            include_conf = st.checkbox("Confirmadas", value=False, key="med_include_conf")

        # Reverse maps for filtering
        obra_name_to_id = {v: k for k, v in obra_id_to_clave.items()}
        trab_name_to_id = {v: k for k, v in trab_id_to_name.items()}

        # Apply filters
        filtered = []
        for c in cabeceras_raw:
            estado = c.get("estado", "Borrador")
            if not include_conf and estado != "Borrador":
                continue

            if obra_filter != "(Todas)":
                if c.get("obra_id") != obra_name_to_id.get(obra_filter):
                    continue

            if trab_filter != "(Todos)":
                if c.get("trabajador_id") != trab_name_to_id.get(trab_filter):
                    continue

            if fecha_filter:
                fecha_c = c.get("fecha", "")
                try:
                    fecha_c = datetime.strptime(str(fecha_c)[:10], "%Y-%m-%d").date()
                    if fecha_c < fecha_filter:
                        continue
                except (ValueError, AttributeError):
                    pass

            filtered.append(c)

        # Sort by date descending (most recent first)
        def _sort_fecha(c):
            f = c.get("fecha", "")
            return str(f)[:10] if f else ""
        filtered.sort(key=_sort_fecha, reverse=True)

        if not filtered:
            st.info("No hay mediciones que coincidan con los filtros.")
        else:
            opts = {}
            for c in filtered:
                obra_id  = c.get("obra_id")
                trab_id  = c.get("trabajador_id")
                obra_nm  = obra_id_to_clave.get(obra_id, str(obra_id))
                trab_nm  = trab_id_to_name.get(trab_id, str(trab_id))
                fecha_nm = str(c.get("fecha", ""))[:10]
                estado   = c.get("estado", "")
                label    = f"{obra_nm}  —  {trab_nm}  —  {fecha_nm}  [{estado}]"
                opts[label] = c["id"]

            sel_label = st.selectbox(
                f"Mediciones encontradas ({len(filtered)})",
                list(opts.keys()), key="med_load_sel",
            )
            col1, col2 = st.columns([1, 5])
            with col1:
                if st.button("📂 Cargar", key="med_btn_load"):
                    sel_id = opts[sel_label]
                    sel_cab = next(c for c in cabeceras_raw if c["id"] == sel_id)
                    lineas_cab = [
                        l for l in all_lineas_raw
                        if l.get("cabecera_id") == sel_id
                    ]
                    _load_borrador_into_state(sel_cab, lineas_cab)
                    st.rerun()
            with col2:
                if st.button("➕ Nueva medición", key="med_btn_nueva"):
                    _reset_session_state()
                    st.rerun()


# ---------------------------------------------------------------------------
# UI — header form
# ---------------------------------------------------------------------------

def _render_header_form(obras_raw: list[dict], trabajadores_raw: list[dict]):
    cab   = st.session_state["med_cabecera"]
    locked = cab.get("estado") == "Confirmado"

    estado = cab.get("estado", "Borrador")
    pill_cls = "med-pill-confirmado" if estado == "Confirmado" else "med-pill-borrador"
    st.markdown(
        f'<span class="js-pill {pill_cls}">{estado}</span>',
        unsafe_allow_html=True,
    )

    # Cascade: Estado → Categoría → Obra
    estados_disponibles = sorted({r.get("estado_obra", "") for r in obras_raw if r.get("estado_obra")})
    c1, c2, c3, c4 = st.columns([1, 1, 2, 1])

    with c1:
        estado_sel = st.selectbox(
            "Estado de obra", ["Todos"] + estados_disponibles,
            key="med_filtro_estado", disabled=locked,
        )
    obras_filtradas = [
        r for r in obras_raw
        if estado_sel == "Todos" or r.get("estado_obra") == estado_sel
    ]

    with c2:
        cats = sorted({r.get("categoria_obra", "") for r in obras_filtradas if r.get("categoria_obra")})
        cat_sel = st.selectbox(
            "Categoría", ["Todas"] + cats,
            key="med_filtro_cat", disabled=locked,
        )
    obras_filtradas = [
        r for r in obras_filtradas
        if cat_sel == "Todas" or r.get("categoria_obra") == cat_sel
    ]

    # Build obra options
    obra_id_to_clave = {r["id"]: r.get("clave", r.get("nombre", str(r["id"]))) for r in obras_filtradas}
    obra_options = sorted(obra_id_to_clave.values())
    obra_clave_to_id = {v: k for k, v in obra_id_to_clave.items()}

    current_obra_id = cab.get("obra_id")
    current_obra_clave = obra_id_to_clave.get(current_obra_id, "")

    default_obra_idx = 0
    if current_obra_clave in obra_options:
        default_obra_idx = obra_options.index(current_obra_clave)

    with c3:
        obra_sel = st.selectbox(
            "Obra *", obra_options if obra_options else ["Sin obras"],
            index=default_obra_idx,
            key="med_obra_sel", disabled=locked,
        )
    selected_obra_id = obra_clave_to_id.get(obra_sel)
    if selected_obra_id != cab.get("obra_id"):
        cab["obra_id"] = selected_obra_id
        _mark_dirty()

    # Subcontratista
    trab_options = sorted([r.get("nombre_completo", "") for r in trabajadores_raw if r.get("nombre_completo")])
    trab_id_to_name = _id_to_name(trabajadores_raw, "nombre_completo")
    trab_name_to_id = _name_to_id(trabajadores_raw, "nombre_completo")

    current_trab_name = trab_id_to_name.get(cab.get("trabajador_id"), "")
    default_trab_idx = 0
    if current_trab_name in trab_options:
        default_trab_idx = trab_options.index(current_trab_name)

    with c4:
        trab_sel = st.selectbox(
            "Subcontratista *", trab_options if trab_options else ["—"],
            index=default_trab_idx,
            key="med_trab_sel", disabled=locked,
        )
    selected_trab_id = trab_name_to_id.get(trab_sel)
    if selected_trab_id != cab.get("trabajador_id"):
        cab["trabajador_id"] = selected_trab_id
        _mark_dirty()

    # Fecha + Observaciones
    col_fecha, col_obs = st.columns([1, 3])
    with col_fecha:
        current_fecha = cab.get("fecha")
        if isinstance(current_fecha, str) and current_fecha:
            try:
                current_fecha = datetime.strptime(current_fecha[:10], "%Y-%m-%d").date()
            except ValueError:
                current_fecha = date.today()
        elif not current_fecha:
            current_fecha = date.today()

        fecha_input = st.date_input(
            "Fecha de medición *", value=current_fecha,
            key="med_fecha_input", disabled=locked,
        )
        if fecha_input != cab.get("fecha"):
            cab["fecha"] = fecha_input
            _mark_dirty()

    with col_obs:
        obs_input = st.text_input(
            "Observaciones (opcional)", value=cab.get("observaciones", ""),
            key="med_obs_input", disabled=locked,
        )
        if obs_input != cab.get("observaciones"):
            cab["observaciones"] = obs_input
            _mark_dirty()

    st.session_state["med_cabecera"] = cab


# ---------------------------------------------------------------------------
# UI — lines editor
# ---------------------------------------------------------------------------

def _render_lineas_editor(
    sectores_raw: list[dict],
    rubros_raw: list[dict],
    obra_id: int | None,
    locked: bool,
):
    st.markdown("##### Líneas de medición")

    # Build options filtered by obra
    filtered_secs  = _sectores_for_obra(obra_id, sectores_raw) if obra_id else []
    sector_options = sorted([s.get("nombre_sector", "") for s in filtered_secs if s.get("nombre_sector")])
    rubro_options  = sorted([r.get("rubro", "") for r in rubros_raw if r.get("rubro")])

    sec_id_name = _id_to_name(filtered_secs, "nombre_sector")
    rub_id_name = _id_to_name(rubros_raw, "rubro")
    sec_name_id = _name_to_id(filtered_secs, "nombre_sector")
    rub_name_id = _name_to_id(rubros_raw, "rubro")

    df = _lineas_to_df(st.session_state["med_lineas"], sec_id_name, rub_id_name)

    column_config = {
        "Sector":         st.column_config.SelectboxColumn("Sector *",        options=sector_options, required=True),
        "Rubro":          st.column_config.SelectboxColumn("Rubro *",          options=rubro_options,  required=True),
        "Descripcion":    st.column_config.TextColumn("Descripción *",         max_chars=200),
        "Unidad":         st.column_config.SelectboxColumn("Unidad",           options=_UNIDAD_OPTIONS),
        "Largo":          st.column_config.NumberColumn("Largo",               min_value=0, format="%.3f"),
        "Ancho":          st.column_config.NumberColumn("Ancho",               min_value=0, format="%.3f"),
        "Alto":           st.column_config.NumberColumn("Alto",                min_value=0, format="%.3f"),
        "Cantidad":       st.column_config.NumberColumn("Cantidad *",          min_value=0, format="%.3f"),
        "PrecioUnitario": st.column_config.NumberColumn("P.U. (Gs.)",          min_value=0, format="%d"),
        "Total":          st.column_config.NumberColumn("Total (Gs.)",         format="%d", disabled=True),
    }

    edited_df = st.data_editor(
        df,
        column_config=column_config,
        num_rows="dynamic" if not locked else "fixed",
        use_container_width=True,
        hide_index=True,
        key="med_data_editor",
        disabled=locked,
    )

    if not locked:
        edited_df = edited_df.reset_index(drop=True)
        new_lineas, new_deleted = _df_to_lineas(
            edited_df,
            st.session_state["med_lineas"],
            sec_name_id,
            rub_name_id,
            st.session_state["med_deleted_ids"],
        )
        if new_lineas != st.session_state["med_lineas"] or new_deleted != st.session_state["med_deleted_ids"]:
            st.session_state["med_lineas"]      = new_lineas
            st.session_state["med_deleted_ids"] = new_deleted
            _mark_dirty()

    return sec_id_name, rub_id_name


# ---------------------------------------------------------------------------
# UI — summary cards
# ---------------------------------------------------------------------------

def _render_summary(lineas: list[dict]):
    if not lineas:
        return
    total_gs = sum(
        float(l.get("cantidad") or 0) * float(l.get("precio_unitario") or 0)
        for l in lineas
    )
    n_sin_precio = sum(1 for l in lineas if not float(l.get("precio_unitario") or 0))
    rubros_usados = len({l.get("rubro_id") for l in lineas if l.get("rubro_id")})

    c1, c2, c3 = st.columns(3)
    c1.metric("Líneas", len(lineas))
    c2.metric("Rubros involucrados", rubros_usados)
    c3.metric("Total estimado", f"₲ {int(total_gs):,}".replace(",", "."))

    if n_sin_precio:
        st.markdown(
            f'<span class="med-sin-precio">⚠ {n_sin_precio} ítem(s) sin precio definido — '
            f'el total es parcial.</span>',
            unsafe_allow_html=True,
        )


# ---------------------------------------------------------------------------
# UI — action bar
# ---------------------------------------------------------------------------

def _render_action_bar(cab_errors: list[str], linea_errors: list[str], cabeceras_raw: list[dict]) -> str:
    """Renderiza botones de acción. Retorna 'guardar' | 'confirmar' | 'reabrir' | ''."""
    estado = st.session_state["med_cabecera"].get("estado", "Borrador")
    locked = estado == "Confirmado"
    all_errors = cab_errors + linea_errors
    action = ""

    st.divider()
    col_g, col_c, col_r, col_status = st.columns([1, 1, 1, 3])

    with col_g:
        if not locked:
            if st.button("💾 Guardar borrador", key="med_btn_guardar", use_container_width=True):
                action = "guardar"

    with col_c:
        if not locked:
            confirm_disabled = bool(all_errors)
            if st.button(
                "✅ Confirmar", key="med_btn_confirmar",
                use_container_width=True, disabled=confirm_disabled,
                type="primary",
            ):
                action = "confirmar"

    with col_r:
        if locked:
            if st.button("🔓 Reabrir", key="med_btn_reabrir", use_container_width=True):
                action = "reabrir"

    with col_status:
        if st.session_state["med_dirty"]:
            st.markdown('<span class="med-dirty">● Cambios sin guardar</span>', unsafe_allow_html=True)
        elif st.session_state["med_saved_at"]:
            ts = st.session_state["med_saved_at"].strftime("%d/%m/%Y %H:%M")
            st.markdown(f'<span class="med-saved">✓ Guardado: {ts}</span>', unsafe_allow_html=True)

    # Validation errors (only if attempted confirm)
    if all_errors and not locked:
        with st.expander("⚠ Errores de validación", expanded=False):
            for e in all_errors:
                st.warning(e)

    return action


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

st.set_page_config(layout="wide", page_title="Mediciones — JS Tools")
st.markdown(_CSS, unsafe_allow_html=True)

_init_session_state()

# Load data
try:
    obras_raw        = get_all_records("dim_obra", ["nombre", "clave", "estado_obra", "categoria_obra"])
    sectores_raw     = get_all_records("dim_sector", ["nombre_sector", "obra_id"])
    rubros_raw       = get_all_records("dim_rubro", ["rubro", "nombre_completo"])
    trabajadores_raw = get_all_records("dim_trabajador", ["nombre_completo"])
    cabeceras_raw    = get_all_records(
        "op_medicion_cabecera",
        ["medicion_ref", "obra_id", "trabajador_id", "fecha", "estado", "observaciones"],
    )
    all_lineas_raw   = get_all_records(
        "op_medicion_linea",
        ["descripcion", "cabecera_id", "sector_id", "rubro_id",
         "unidad", "largo", "ancho", "alto", "cantidad", "precio_unitario"],
    )
except ConnectionError as e:
    st.error(f"Error de conexión con Supabase: {e}")
    st.stop()

# Build name maps (integer keys)
obra_id_to_clave  = {r["id"]: r.get("clave", r.get("nombre", str(r["id"]))) for r in obras_raw}
trab_id_to_name   = {r["id"]: r.get("nombre_completo", "") for r in trabajadores_raw}
sec_id_name_full  = {r["id"]: r.get("nombre_sector", "") for r in sectores_raw}
rub_id_name_full  = {r["id"]: r.get("rubro", "") for r in rubros_raw}

# Page title
st.title("📐 Mediciones de Obra")

tab_editar, tab_general = st.tabs(["✏️ Editar", "📊 Vista General"])

# ── Tab 1: Editor de mediciones ──────────────────────────────────────────────
with tab_editar:
    st.markdown(
        '<p class="js-sub">Planilla de cómputo métrico semanal por subcontratista.</p>',
        unsafe_allow_html=True,
    )

    # Load section
    _render_load_section(cabeceras_raw, obra_id_to_clave, trab_id_to_name, all_lineas_raw)

    st.divider()

    # Header form
    _render_header_form(obras_raw, trabajadores_raw)

    cab        = st.session_state["med_cabecera"]
    obra_id    = cab.get("obra_id")
    locked     = cab.get("estado") == "Confirmado"

    st.divider()

    # Lines editor
    sec_id_name, rub_id_name = _render_lineas_editor(sectores_raw, rubros_raw, obra_id, locked)

    st.divider()

    # Summary
    _render_summary(st.session_state["med_lineas"])

    # Validate
    cab_errors       = _validate_cabecera(cab)
    linea_errors, n_sin_precio = _validate_lineas(st.session_state["med_lineas"])

    # Action bar
    action = _render_action_bar(cab_errors, linea_errors, cabeceras_raw)

    # Duplicate warning before confirming
    if action == "confirmar":
        fecha_val = cab.get("fecha")
        if isinstance(fecha_val, str):
            try:
                fecha_val = datetime.strptime(fecha_val[:10], "%Y-%m-%d").date()
            except ValueError:
                fecha_val = None
        is_dup = _check_duplicate_confirmada(
            cab.get("obra_id"), cab.get("trabajador_id"), fecha_val,
            cabeceras_raw, st.session_state["med_cabecera_id"],
        )
        if is_dup:
            st.warning(
                "⚠ Ya existe una medición **Confirmada** para esta obra y subcontratista "
                "en la misma semana. ¿Querés confirmar de todas formas?"
            )
            if st.button("Sí, confirmar igual", key="med_btn_confirm_dup"):
                if _confirmar(cabeceras_raw):
                    st.toast("Medición confirmada.", icon="✅")
                    st.rerun()
        else:
            if _confirmar(cabeceras_raw):
                st.toast("Medición confirmada.", icon="✅")
                st.rerun()

    elif action == "guardar":
        if _guardar(cabeceras_raw):
            st.toast("Borrador guardado.", icon="💾")
            st.rerun()

    elif action == "reabrir":
        if _reabrir():
            st.toast("Medición reabierta.", icon="🔓")
            st.rerun()

    # Export section
    if st.session_state["med_cabecera_id"] and st.session_state["med_lineas"]:
        st.divider()
        st.markdown("##### Exportar")

        obra_id_exp  = cab.get("obra_id")
        trab_id_exp  = cab.get("trabajador_id")
        obra_nombre  = obra_id_to_clave.get(obra_id_exp, str(obra_id_exp))
        trab_nombre  = trab_id_to_name.get(trab_id_exp, str(trab_id_exp))

        fecha_val = cab.get("fecha")
        if isinstance(fecha_val, date):
            fecha_str = fecha_val.strftime("%d/%m/%Y")
        elif isinstance(fecha_val, str):
            fecha_str = fecha_val[:10]
        else:
            fecha_str = ""

        try:
            today_str = date.today().strftime("%d/%m/%Y")
        except Exception:
            today_str = str(date.today())

        # Enrich lineas for PDF (add _sector_name, _rubro_name)
        lineas_enrich = []
        for l in st.session_state["med_lineas"]:
            le = dict(l)
            le["_sector_name"] = sec_id_name.get(l.get("sector_id"), "—")
            le["_rubro_name"]  = rub_id_name.get(l.get("rubro_id"), "—")
            lineas_enrich.append(le)

        nombre_archivo = f"Medicion - {obra_nombre} - {trab_nombre}"

        col_pdf, col_xls = st.columns(2)
        with col_pdf:
            try:
                pdf_bytes = generate_medicion_pdf(
                    obra_nombre=obra_nombre,
                    subcontratista_nombre=trab_nombre,
                    fecha_medicion=fecha_str,
                    observaciones=cab.get("observaciones", ""),
                    lineas=lineas_enrich,
                    fecha_generacion=today_str,
                )
                st.download_button(
                    "📄 Descargar PDF",
                    data=pdf_bytes,
                    file_name=f"{nombre_archivo}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Error al generar PDF: {e}")

        with col_xls:
            try:
                xlsx_bytes = _generate_excel(
                    cab, st.session_state["med_lineas"],
                    sec_id_name, rub_id_name,
                    obra_nombre, trab_nombre,
                )
                st.download_button(
                    "📊 Descargar Excel",
                    data=xlsx_bytes,
                    file_name=f"{nombre_archivo}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Error al generar Excel: {e}")

# ── Tab 2: Vista General (acumulado) ────────────────────────────────────────
with tab_general:
    _render_vista_general(
        obras_raw, trabajadores_raw, cabeceras_raw, all_lineas_raw,
        sec_id_name_full, rub_id_name_full,
        obra_id_to_clave, trab_id_to_name,
    )
