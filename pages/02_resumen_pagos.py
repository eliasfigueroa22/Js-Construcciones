"""Resumen de Pagos — ficha técnica de pagos por trabajador y obra.

Flujo:
1. El usuario selecciona trabajador y obra (opciones cargadas dinámicamente).
2. Presiona "Generar Reporte" para cruzar fact_pago con fact_presupuesto_subcontratista.
3. Se muestra la ficha con tarjetas de resumen y tabla de detalle.
4. Se ofrece descarga en PDF con bloque de resumen + tabla de detalle.
"""

import base64
from datetime import datetime
from io import BytesIO

import streamlit.components.v1 as components
from streamlit_sortables import sort_items

import pandas as pd
import streamlit as st
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from connectors.supabase_connector import get_all_records
from core.base_tool import ToolMetadata

# -- Metadatos (leídos por registry.py vía AST, no ejecutados) ----------------
TOOL = ToolMetadata(
    name="Resumen de Pagos",
    description="Generá fichas técnicas de pagos por trabajador y obra, con comparativa contra presupuesto.",
    icon="💰",
    page_file="02_resumen_pagos.py",
)

# -- Estilos PDF ---------------------------------------------------------------
# Prefijo "Pagos" para evitar conflictos con los estilos de pdf_generator.py.
_STYLES = getSampleStyleSheet()

_PDF_CELL = ParagraphStyle(
    "PagosCell", parent=_STYLES["Normal"], fontSize=7, leading=9
)
_PDF_CELL_BOLD = ParagraphStyle(
    "PagosCellBold", parent=_STYLES["Normal"], fontSize=7, leading=9,
    fontName="Helvetica-Bold",
)
_PDF_HEADER = ParagraphStyle(
    "PagosHeader", parent=_STYLES["Normal"], fontSize=8, leading=10,
    fontName="Helvetica-Bold", textColor=colors.white,
)
_PDF_CARD_LABEL = ParagraphStyle(
    "PagosCardLabel", parent=_STYLES["Normal"], fontSize=7, leading=9,
    fontName="Helvetica-Bold", textColor=colors.white,
)
_PDF_CARD_VALUE = ParagraphStyle(
    "PagosCardValue", parent=_STYLES["Normal"], fontSize=10, leading=12,
    fontName="Helvetica-Bold", textColor=colors.HexColor("#2C3E50"),
)
_PDF_CARD_NOTE = ParagraphStyle(
    "PagosCardNote", parent=_STYLES["Normal"], fontSize=7, leading=9,
    textColor=colors.HexColor("#7f8c8d"),
)

# Columnas del reporte de detalle (usadas en la tabla y en el PDF).
# Dos columnas de dinero separadas: una para pagos, otra para presupuestos.
_DETAIL_COLUMNS = [
    {"key": "fecha",       "label": "Fecha",              "width_mm": 20, "align": "CENTER"},
    {"key": "concepto",    "label": "Concepto",           "width_mm": 60, "align": "LEFT"},
    {"key": "tipo",        "label": "Tipo / Estado",      "width_mm": 22, "align": "CENTER"},
    {"key": "monto_pres",  "label": "Presupuesto (Gs.)",  "width_mm": 28, "align": "RIGHT", "is_money": True},
    {"key": "monto_pago",  "label": "Pago (Gs.)",         "width_mm": 28, "align": "RIGHT", "is_money": True},
    {"key": "metodo",      "label": "Método",             "width_mm": 22, "align": "CENTER"},
]


# -- Helpers -------------------------------------------------------------------
def _fmt_gs(value) -> str:
    """Formatea un número al estilo paraguayo: 1.234.567."""
    try:
        return f"{int(float(value)):,}".replace(",", ".")
    except (ValueError, TypeError):
        return str(value)


def _fmt_date(val) -> str:
    """Convierte una fecha ISO (YYYY-MM-DD) a DD/MM/YYYY."""
    if not val:
        return "—"
    try:
        return datetime.strptime(str(val)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return str(val)


def _safe_xml(text) -> str:
    """Escapa caracteres especiales XML para Paragraph de reportlab."""
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _generate_excel(
    trabajador: str,
    obra: str,
    today: str,
    presup_records: list[dict],
    pago_records: list[dict],
    total_pagado: float,
    total_presupuestado: float | None,
) -> bytes:
    """Genera el Excel de la ficha técnica con openpyxl."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Detalle"

    DARK      = "1B4F72"
    GREY      = "D5DBDB"
    LIGHT     = "EBF5FB"
    TEXT_DARK = "2C3E50"
    thin = Side(border_style="thin", color="CCCCCC")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _st(cell, bold=False, bg=None, fg="000000", halign="center",
            size=10, italic=False, num_fmt=None):
        cell.font = Font(bold=bold, color=fg, size=size, italic=italic)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=halign, vertical="center")
        if num_fmt:
            cell.number_format = num_fmt

    # --- Filas de encabezado ---
    ws.append([f"Ficha Técnica — {trabajador}"])
    _st(ws["A1"], bold=True, fg=DARK, size=14)
    ws.row_dimensions[1].height = 22

    ws.append([f"Obra: {obra}"])
    _st(ws["A2"], fg=TEXT_DARK, size=11)

    ws.append([f"Generado el {today}"])
    _st(ws["A3"], italic=True, fg="7F8C8D", size=9)

    ws.append([])  # fila 4: separador

    # --- Bloque resumen (fila 5-6) ---
    saldo = (total_presupuestado - total_pagado) if total_presupuestado is not None else None
    lbl_row = 5
    val_row = 6
    summary = [
        ("Presupuesto Total", f"Gs. {_fmt_gs(total_presupuestado)}" if total_presupuestado is not None else "N/D"),
        ("Total Pagado",      f"Gs. {_fmt_gs(total_pagado)}"),
        ("Saldo",             f"Gs. {_fmt_gs(saldo)}" if saldo is not None else "N/D"),
    ]
    for idx, (lbl, val) in enumerate(summary):
        col = idx * 2 + 1  # columnas 1, 3, 5
        ws.cell(lbl_row, col, lbl)
        _st(ws.cell(lbl_row, col), bold=True, bg=DARK, fg="FFFFFF", size=9)
        ws.cell(val_row, col, val)
        _st(ws.cell(val_row, col), bold=True, fg=TEXT_DARK, size=11)

    ws.row_dimensions[lbl_row].height = 16
    ws.row_dimensions[val_row].height = 20
    ws.append([])  # fila 7: separador

    # --- Cabecera de tabla ---
    headers = ["Fecha", "Concepto", "Tipo / Estado", "Presupuesto (Gs.)", "Pago (Gs.)", "Método"]
    ws.append(headers)
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        _st(cell, bold=True, bg=DARK, fg="FFFFFF", size=10)
    ws.row_dimensions[hdr_row].height = 18

    # Freeze + auto-filter desde la cabecera
    ws.freeze_panes = f"A{hdr_row + 1}"
    ws.auto_filter.ref = f"A{hdr_row}:F{hdr_row}"

    # --- Filas de datos ---
    for r in presup_records:
        ws.append([
            _fmt_date(r.get("fecha_presupuesto", "")) or "—",
            r.get("concepto") or f"Presupuesto N°{r.get('presupuesto_nro', '?')}",
            r.get("estado", "") or "",
            float(r.get("monto_presupuestado", 0) or 0),
            None,
            "",
        ])

    for r in pago_records:
        ws.append([
            _fmt_date(r.get("fecha_pago", "")),
            r.get("concepto", "") or "",
            r.get("tipo_pago", "") or "",
            None,
            float(r.get("monto_pago", 0) or 0),
            r.get("metodo_pago", "") or "",
        ])

    # --- Fila TOTAL ---
    ws.append([
        "TOTAL", "", "",
        total_presupuestado if total_presupuestado is not None else None,
        total_pagado,
        "",
    ])
    tot_row = ws.max_row

    # Aplicar estilos a todas las filas de datos + total
    for i, row in enumerate(ws.iter_rows(min_row=hdr_row + 1, max_row=tot_row)):
        is_total = (row[0].row == tot_row)
        for cell in row:
            col = cell.column
            cell.border = border_all
            # Alineación
            if col == 2:
                cell.alignment = Alignment(horizontal="left",   vertical="center")
            elif col in (4, 5):
                cell.alignment = Alignment(horizontal="right",  vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Fondo
            if is_total:
                _st(cell, bold=True, bg=GREY, halign=cell.alignment.horizontal or "center")
            elif i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
            # Formato numérico para columnas de dinero
            if col in (4, 5) and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    # Fila de total: borde superior más grueso
    med = Side(border_style="medium", color=DARK)
    for cell in ws[tot_row]:
        cell.border = Border(left=thin, right=thin, bottom=thin, top=med)

    # Anchos de columna
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# -- Generador PDF -------------------------------------------------------------
def _generate_pagos_pdf(
    trabajador: str,
    obra: str,
    today: str,
    total_presupuestado: float | None,
    total_pagado: float,
    balance: float | None,
    rows: list[dict],
) -> bytes:
    """Genera el PDF del resumen con bloque de tarjetas y tabla de detalle."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=30 * mm,
        rightMargin=20 * mm,
        topMargin=25 * mm,
        bottomMargin=25 * mm,
    )
    available_width = A4[0] - 50 * mm
    story: list = []

    # Título y subtítulo
    story.append(
        Paragraph(_safe_xml(f"Resumen de Pagos — {trabajador}"), _STYLES["Title"])
    )
    story.append(
        Paragraph(_safe_xml(f"{obra}  |  Generado el {today}"), _STYLES["Normal"])
    )
    story.append(Spacer(1, 6 * mm))

    # -- Bloque de resumen: 3 tarjetas lado a lado ----------------------------
    presup_text = f"Gs. {_fmt_gs(total_presupuestado)}" if total_presupuestado is not None else "N/D"
    pagado_text = f"Gs. {_fmt_gs(total_pagado)}"
    balance_text = f"Gs. {_fmt_gs(balance)}" if balance is not None else "N/D"

    pct_note = ""
    if total_presupuestado:
        pct_note = f"{total_pagado / total_presupuestado * 100:.1f}% ejecutado"
    balance_note = ""
    if balance is not None:
        balance_note = "A favor" if balance >= 0 else "En contra"

    card_data = [
        [
            Paragraph("<b>PRESUPUESTO TOTAL</b>", _PDF_CARD_LABEL),
            Paragraph("<b>TOTAL PAGADO</b>", _PDF_CARD_LABEL),
            Paragraph("<b>SALDO</b>", _PDF_CARD_LABEL),
        ],
        [
            Paragraph(_safe_xml(presup_text), _PDF_CARD_VALUE),
            Paragraph(_safe_xml(pagado_text), _PDF_CARD_VALUE),
            Paragraph(_safe_xml(balance_text), _PDF_CARD_VALUE),
        ],
        [
            Paragraph("", _PDF_CARD_NOTE),
            Paragraph(_safe_xml(pct_note), _PDF_CARD_NOTE),
            Paragraph(_safe_xml(balance_note), _PDF_CARD_NOTE),
        ],
    ]

    card_w = available_width / 3
    card_table = Table(card_data, colWidths=[card_w] * 3)
    card_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4F72")),
        ("BACKGROUND", (0, 1), (-1, 2), colors.HexColor("#F8F9FA")),
        ("BOX",        (0, 0), (-1, -1), 1,   colors.HexColor("#1B4F72")),
        ("INNERGRID",  (0, 0), (-1, -1), 0.5, colors.HexColor("#d5dbdb")),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(card_table)
    story.append(Spacer(1, 8 * mm))

    # -- Tabla de detalle de pagos --------------------------------------------
    col_widths = [c["width_mm"] * mm for c in _DETAIL_COLUMNS]
    total_w = sum(col_widths)
    if total_w > available_width:
        scale = available_width / total_w
        col_widths = [w * scale for w in col_widths]

    header_row = [Paragraph(_safe_xml(c["label"]), _PDF_HEADER) for c in _DETAIL_COLUMNS]
    table_data: list = [header_row]

    for row in rows:
        data_row = []
        for col in _DETAIL_COLUMNS:
            val = row.get(col["key"], "")
            if col.get("is_money"):
                num = float(val or 0)
                val = _fmt_gs(num) if num != 0 else "—"
            data_row.append(Paragraph(_safe_xml(str(val)), _PDF_CELL))
        table_data.append(data_row)

    # Fila TOTAL — suma independiente por cada columna de dinero
    money_totals = {
        col["key"]: sum(float(r.get(col["key"], 0) or 0) for r in rows)
        for col in _DETAIL_COLUMNS
        if col.get("is_money")
    }
    totals_row = []
    placed_label = False
    for col in _DETAIL_COLUMNS:
        if col.get("is_money"):
            totals_row.append(
                Paragraph(_safe_xml(_fmt_gs(money_totals[col["key"]])), _PDF_CELL_BOLD)
            )
        elif not placed_label:
            totals_row.append(Paragraph("<b>TOTAL</b>", _PDF_CELL_BOLD))
            placed_label = True
        else:
            totals_row.append(Paragraph("", _PDF_CELL))
    table_data.append(totals_row)

    detail_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#2c3e50")),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  6),
        ("TOPPADDING",    (0, 0), (-1, 0),  4),
        ("TOPPADDING",    (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]
    # Alineación por columna
    for idx, col in enumerate(_DETAIL_COLUMNS):
        align = col.get("align", "LEFT")
        style_cmds.append(("ALIGN", (idx, 1), (idx, -1), align))

    # Filas alternadas
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style_cmds.append(
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#ecf0f1"))
            )

    # Estilo de la fila de totales
    totals_idx = len(table_data) - 1
    style_cmds.append(
        ("BACKGROUND", (0, totals_idx), (-1, totals_idx), colors.HexColor("#d5dbdb"))
    )
    style_cmds.append(
        ("LINEABOVE", (0, totals_idx), (-1, totals_idx), 1.5, colors.HexColor("#2c3e50"))
    )

    detail_table.setStyle(TableStyle(style_cmds))
    story.append(detail_table)

    doc.build(story)
    return buffer.getvalue()


# -- CSS ----------------------------------------------------------------------
st.markdown("""
<style>
:root {
    --js-accent:  #E8622A;
    --js-success: #27AE60;
    --js-danger:  #E74C3C;
    --js-warn:    #E67E22;
    --js-muted:   #6B7280;
    --js-border:  rgba(255,255,255,0.09);
    --js-surface: rgba(255,255,255,0.03);
}
.js-sub {
    color: var(--js-muted);
    font-size: .875rem;
    margin-top: -10px;
    margin-bottom: 24px;
}
.js-report-header {
    padding: 16px 20px;
    border-left: 4px solid var(--js-accent);
    border-radius: 4px;
    background: rgba(255,255,255,0.04);
    margin-bottom: 8px;
}
.js-report-title {
    margin: 0;
    font-size: 1.3rem;
    font-weight: 700;
}
.js-report-sub {
    margin: 4px 0 0;
    font-size: 0.95rem;
}
.js-report-date {
    font-size: 0.78rem;
    color: var(--js-muted);
    margin-top: 4px;
}
</style>
""", unsafe_allow_html=True)

# -- UI principal --------------------------------------------------------------
st.title("💰 Resumen de Pagos")
st.markdown(
    '<p class="js-sub">Ficha técnica de pagos por trabajador y obra · Comparativa contra presupuesto</p>',
    unsafe_allow_html=True,
)

# == Paso 1: Filtros ===========================================================
st.subheader("1. Filtros")

try:
    with st.spinner("Cargando registros..."):
        all_pagos  = get_all_records("fact_pago")
        obras_list = get_all_records("dim_obra")
        trab_recs  = get_all_records("dim_trabajador")
        nm_trab = {r["id"]: r.get("nombre_completo", str(r["id"])) for r in trab_recs}
        nm_obra = {r["id"]: r.get("clave") or r.get("nombre", str(r["id"])) for r in obras_list}
except ConnectionError:
    st.error("No se pudo conectar con la base de datos. Verificá tu conexión.")
    st.stop()
except Exception as exc:
    st.error(f"Error inesperado al cargar datos: {exc}")
    st.stop()

if not all_pagos:
    st.info("No se encontraron registros de pagos.")
    st.stop()

# Lookup: nombre de obra → atributos (estado_obra, categoria_obra)
obra_attrs: dict[str, dict] = {
    r.get("clave") or r.get("nombre", str(r["id"])): r
    for r in obras_list
}

# Nombres de obras que efectivamente tienen pagos registrados
pagos_obra_names: set[str] = {
    nm_obra.get(r.get("obra_id"), "")
    for r in all_pagos
    if r.get("obra_id") is not None
}
pagos_obra_names.discard("")

# -- Filtros jerárquicos (4 niveles) ------------------------------------------
fcol1, fcol2 = st.columns(2)

# Nivel 1 — Estado de la obra
estadoes = sorted({
    obra_attrs[n].get("estado_obra") or "Sin estado"
    for n in pagos_obra_names
    if n in obra_attrs
})
with fcol1:
    selected_estado = st.selectbox("Estado de la obra", estadoes, key="pagos_estado_select")

# Nivel 2 — Categoría / Tipo de obra (filtrado por estado)
categorias = sorted({
    obra_attrs[n].get("categoria_obra") or "Sin categoría"
    for n in pagos_obra_names
    if n in obra_attrs
    and (obra_attrs[n].get("estado_obra") or "Sin estado") == selected_estado
})
with fcol2:
    selected_categoria = st.selectbox("Tipo de obra", categorias, key="pagos_categoria_select")

# Nivel 3 — Nombre de la obra (filtrado por estado + categoría)
obras_filtradas = sorted({
    n for n in pagos_obra_names
    if n in obra_attrs
    and (obra_attrs[n].get("estado_obra") or "Sin estado") == selected_estado
    and (obra_attrs[n].get("categoria_obra") or "Sin categoría") == selected_categoria
})
obra_options = ["Todas las obras"] + list(obras_filtradas)

fcol3, fcol4 = st.columns(2)
with fcol3:
    selected_obra = st.selectbox("Obra", obra_options, key="pagos_obra_select")

# Nivel 4 — Trabajador (filtrado por la obra seleccionada)
def _obra_en_scope_filter(obra_id) -> bool:
    nombre = nm_obra.get(obra_id, "")
    if selected_obra != "Todas las obras":
        return nombre == selected_obra
    attrs = obra_attrs.get(nombre, {})
    return (
        (attrs.get("estado_obra") or "Sin estado") == selected_estado
        and (attrs.get("categoria_obra") or "Sin categoría") == selected_categoria
    )

trabajadores_en_scope = sorted({
    nm_trab.get(r.get("trabajador_id"), "")
    for r in all_pagos
    if r.get("trabajador_id") is not None
    and r.get("obra_id") is not None
    and _obra_en_scope_filter(r.get("obra_id"))
} - {""})

if not trabajadores_en_scope:
    st.warning("No se encontraron trabajadores para los filtros seleccionados.")
    st.stop()

with fcol4:
    selected_trabajador = st.selectbox(
        "Personal", trabajadores_en_scope, key="pagos_trabajador_select"
    )

if st.button("Generar Reporte", type="primary", use_container_width=True, key="pagos_generar"):
    try:
        with st.spinner("Consultando base de datos..."):
            pagos = [
                r for r in all_pagos
                if nm_trab.get(r.get("trabajador_id"), "") == selected_trabajador
                and r.get("obra_id") is not None
                and _obra_en_scope_filter(r.get("obra_id"))
            ]
            pagos.sort(key=lambda r: r.get("fecha_pago", "") or "")

            # Presupuestos correspondientes (mismos criterios)
            all_presupuestos = get_all_records("fact_presupuesto_subcontratista")
            presupuestos = [
                r for r in all_presupuestos
                if nm_trab.get(r.get("trabajador_id"), "") == selected_trabajador
                and r.get("obra_id") is not None
                and _obra_en_scope_filter(r.get("obra_id"))
            ]

        total_presupuestado = (
            sum(float(r.get("monto_presupuestado", 0) or 0) for r in presupuestos)
            if presupuestos else None
        )
        total_pagado = sum(float(r.get("monto_pago", 0) or 0) for r in pagos)
        balance = (
            (total_presupuestado - total_pagado)
            if total_presupuestado is not None else None
        )

        obra_display = selected_obra if selected_obra != "Todas las obras" else (
            f"Todas las obras ({selected_estado} / {selected_categoria})"
        )

        st.session_state["pagos_data"] = pagos
        st.session_state["pagos_presupuesto_records"] = presupuestos
        st.session_state["pagos_presupuesto"] = total_presupuestado
        st.session_state["pagos_total"] = total_pagado
        st.session_state["pagos_balance"] = balance
        st.session_state["pagos_trabajador"] = selected_trabajador
        st.session_state["pagos_obra"] = obra_display

    except ConnectionError:
        st.error("No se pudo conectar con la base de datos. Verificá tu conexión.")
    except Exception as exc:
        st.error(f"Error inesperado al generar reporte: {exc}")

# == Paso 2: Reporte ===========================================================
pagos_data = st.session_state.get("pagos_data")
if pagos_data is None:
    st.stop()

trabajador_name = st.session_state["pagos_trabajador"]
obra_name       = st.session_state["pagos_obra"]
total_presupuestado = st.session_state["pagos_presupuesto"]
total_pagado        = st.session_state["pagos_total"]
balance             = st.session_state["pagos_balance"]
presupuesto_records = st.session_state.get("pagos_presupuesto_records", [])

today          = datetime.now().strftime("%d/%m/%Y")
today_filename = datetime.now().strftime("%Y%m%d")

if not pagos_data:
    st.info(
        f"No se encontraron pagos para **{trabajador_name}** en **{obra_name}**."
    )
    st.stop()

st.divider()

# -- Encabezado del reporte ---------------------------------------------------
st.markdown(
    f"""
    <div class="js-report-header">
      <div class="js-report-title">Resumen de Pagos</div>
      <div class="js-report-sub">{trabajador_name} — {obra_name}</div>
      <div class="js-report-date">Generado el {today}</div>
    </div>
    """,
    unsafe_allow_html=True,
)

# -- Tarjetas de resumen -------------------------------------------------------
mcol1, mcol2, mcol3 = st.columns(3)

if total_presupuestado is not None:
    mcol1.metric("Presupuesto Total", f"Gs. {_fmt_gs(total_presupuestado)}")
else:
    with mcol1:
        st.warning("Sin presupuesto registrado para esta combinación.")

pct_delta = None
if total_presupuestado:
    pct_delta = f"{total_pagado / total_presupuestado * 100:.1f}% ejecutado"
mcol2.metric("Total Pagado", f"Gs. {_fmt_gs(total_pagado)}", delta=pct_delta)

if balance is not None:
    delta_lbl = "A favor" if balance >= 0 else "En contra"
    delta_clr = "normal" if balance >= 0 else "inverse"
    mcol3.metric(
        "Saldo", f"Gs. {_fmt_gs(balance)}",
        delta=delta_lbl, delta_color=delta_clr,
    )

# -- Tabla de detalle combinada (presupuestos + pagos) -------------------------
st.divider()
st.subheader("Detalle")

presup_display = [
    {
        "Fecha":             _fmt_date(r.get("fecha_presupuesto", "")) or "—",
        "Concepto":          r.get("concepto") or f"Presupuesto N°{r.get('presupuesto_nro', '?')}",
        "Tipo / Estado":     r.get("estado", "—") or "—",
        "Presupuesto (Gs.)": _fmt_gs(r.get("monto_presupuestado", 0)),
        "Pago (Gs.)":        "—",
        "Método":            "—",
    }
    for r in presupuesto_records
]
pago_display = [
    {
        "Fecha":             _fmt_date(r.get("fecha_pago", "")),
        "Concepto":          r.get("concepto", "—") or "—",
        "Tipo / Estado":     r.get("tipo_pago", "—") or "—",
        "Presupuesto (Gs.)": "—",
        "Pago (Gs.)":        _fmt_gs(r.get("monto_pago", 0)),
        "Método":            r.get("metodo_pago", "—") or "—",
    }
    for r in pagos_data
]
display_rows = presup_display + pago_display

_all_detail_labels = [c["label"] for c in _DETAIL_COLUMNS]

if "_col_order_02" not in st.session_state:
    st.session_state["_col_order_02"] = _all_detail_labels.copy()

with st.popover("⚙️ Columnas"):
    for _lbl in _all_detail_labels:
        _checked = _lbl in st.session_state["_col_order_02"]
        if st.checkbox(_lbl, value=_checked, key=f"chk_02_{_lbl}"):
            if not _checked:
                st.session_state["_col_order_02"].append(_lbl)
        else:
            if _checked:
                st.session_state["_col_order_02"].remove(_lbl)

_cur_02 = st.session_state["_col_order_02"]
if _cur_02:
    st.caption("Arrastrá para reordenar")
    _cur_02 = sort_items(_cur_02, key=f"sort_02_{abs(hash(frozenset(_cur_02)))}")
    st.session_state["_col_order_02"] = _cur_02
_visible_labels = _cur_02

_df_detail = pd.DataFrame(display_rows)
_show_cols = [l for l in _visible_labels if l in _df_detail.columns]
if _show_cols:
    _df_detail = _df_detail[_show_cols]

_col_widths = {c["label"]: c["width_mm"] for c in _DETAIL_COLUMNS}
st.dataframe(
    _df_detail,
    use_container_width=True,
    hide_index=True,
    column_config={
        lbl: st.column_config.TextColumn(lbl, width=int(_col_widths[lbl] * 3.78))
        for lbl in _show_cols
        if lbl in _col_widths
    },
)

totals_md = f"Total Pagado: Gs. {_fmt_gs(total_pagado)}"
if total_presupuestado is not None:
    totals_md += f"&nbsp;&nbsp;|&nbsp;&nbsp;Total Presupuestado: Gs. {_fmt_gs(total_presupuestado)}"
st.markdown(
    f"<div style='text-align:right;font-weight:bold;padding:4px 0'>{totals_md}</div>",
    unsafe_allow_html=True,
)

# -- Descarga PDF --------------------------------------------------------------
st.divider()

# Presupuestos primero, luego pagos (misma lógica que la tabla UI)
pdf_rows = [
    {
        "fecha":       _fmt_date(r.get("fecha_presupuesto", "")) or "—",
        "concepto":    r.get("concepto") or f"Presupuesto N°{r.get('presupuesto_nro', '?')}",
        "tipo":        r.get("estado", "") or "",
        "monto_pago":  0,
        "monto_pres":  float(r.get("monto_presupuestado", 0) or 0),
        "metodo":      "",
    }
    for r in presupuesto_records
] + [
    {
        "fecha":       _fmt_date(r.get("fecha_pago", "")),
        "concepto":    r.get("concepto", "") or "",
        "tipo":        r.get("tipo_pago", "") or "",
        "monto_pago":  float(r.get("monto_pago", 0) or 0),
        "monto_pres":  0,
        "metodo":      r.get("metodo_pago", "") or "",
    }
    for r in pagos_data
]

pdf_bytes = _generate_pagos_pdf(
    trabajador=trabajador_name,
    obra=obra_name,
    today=today,
    total_presupuestado=total_presupuestado,
    total_pagado=total_pagado,
    balance=balance,
    rows=pdf_rows,
)

pdf_b64 = base64.b64encode(pdf_bytes).decode("utf-8")

xl_bytes = _generate_excel(
    trabajador=trabajador_name,
    obra=obra_name,
    today=today,
    presup_records=presupuesto_records,
    pago_records=pagos_data,
    total_pagado=total_pagado,
    total_presupuestado=total_presupuestado,
)

bcol1, bcol2, bcol3 = st.columns(3)
with bcol1:
    st.download_button(
        label="📥 Descargar PDF",
        data=pdf_bytes,
        file_name=f"Ficha Técnica - {trabajador_name} - {obra_name}.pdf",
        mime="application/pdf",
        use_container_width=True,
        key="pagos_dl_pdf",
    )
with bcol2:
    st.download_button(
        label="📊 Descargar Excel",
        data=xl_bytes,
        file_name=f"Ficha Técnica - {trabajador_name} - {obra_name}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="pagos_dl_excel",
    )
with bcol3:
    # El botón vive dentro de un components.html para que el click sea un
    # gesto directo del browser (sin rerun de Streamlit), evitando que el
    # popup blocker bloquee window.open(). El PDF se abre en una nueva
    # pestaña y a los 1200ms se dispara el diálogo de impresión.
    components.html(
        f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ padding:3px 0; }}
  button {{
    width:100%; cursor:pointer; display:block;
  }}
</style>
</head>
<body>
<button id="btn">🖨️ Imprimir PDF</button>
<script>
// Copia los estilos computados del botón de descarga nativo (o cualquier
// botón secundario de Streamlit) para que luzca idéntico en cualquier tema.
function syncStyle() {{
  try {{
    var p  = window.parent;
    var ps = p.getComputedStyle(p.document.documentElement);
    var bg = ps.getPropertyValue('--background-color').trim()
           || p.getComputedStyle(p.document.body).backgroundColor;
    if (bg) document.body.style.background = bg;

    // Prioridad: botón de descarga > cualquier botón secundario > primer botón
    var ref = p.document.querySelector('[data-testid="stDownloadButton"] button')
           || p.document.querySelector('[data-testid="baseButton-secondary"]')
           || p.document.querySelector('button');
    if (!ref) return;

    var rs  = p.getComputedStyle(ref);
    var btn = document.getElementById('btn');
    btn.style.color           = rs.color;
    btn.style.backgroundColor = rs.backgroundColor;
    btn.style.borderColor     = rs.borderColor;
    btn.style.borderWidth     = rs.borderWidth;
    btn.style.borderStyle     = rs.borderStyle;
    btn.style.borderRadius    = rs.borderRadius;
    btn.style.fontSize        = rs.fontSize;
    btn.style.fontFamily      = rs.fontFamily;
    btn.style.fontWeight      = rs.fontWeight;
    btn.style.lineHeight      = rs.lineHeight;
    btn.style.padding         = rs.padding;
    btn.style.height          = rs.height;
  }} catch(e) {{}}
}}
syncStyle();
setTimeout(syncStyle, 400);

var PDF_B64 = '{pdf_b64}';
document.getElementById('btn').onclick = function() {{
  var bin = atob(PDF_B64);
  var arr = new Uint8Array(bin.length);
  for (var i = 0; i < bin.length; i++) arr[i] = bin.charCodeAt(i);
  var blob = new Blob([arr], {{type:'application/pdf'}});
  var url  = URL.createObjectURL(blob);
  var win  = window.open(url, '_blank');
  if (win) setTimeout(function() {{ win.print(); }}, 1200);
}};
</script>
</body></html>""",
        height=50,
    )
