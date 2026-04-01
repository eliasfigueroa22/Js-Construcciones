#!/usr/bin/env python3
"""migrate_certs.py — Importar certificados históricos desde Excel a Airtable.

Uso:
    python scripts/migrate_certs.py --raiz "ruta/a/carpeta_raiz"
    python scripts/migrate_certs.py --raiz "ruta/a/carpeta_raiz" --dry-run

La carpeta raíz debe contener subcarpetas, una por obra. Ej:
    raiz/
      ALGESA/       ← archivos ALGESA_cert1.xlsx, ALGESA_cert2.xlsx, ...
      CICSA/        ← archivos CICSA cert 1.xlsx, CICSA cert 2.xlsx, ...

El script:
  1. Carga todas las obras de Airtable.
  2. Por cada subcarpeta, muestra la lista de obras y te pide que elijas cuál corresponde.
  3. Lee todos los .xlsx de la subcarpeta y los ordena por número de certificado.
  4. Del primer archivo extrae el presupuesto base y lo sube a CertPresupuestoLinea.
  5. De cada archivo sube una CertCabecera (estado Confirmado) + las CertLinea con Cant. Actual > 0.

Formatos soportados:
  - Variante A: Item en col A (OFICINAS, COMEDOR, BAÑO en Cert10)
  - Variante B: Item en col B (Garita, cercado, archivos D27)
  - Multi-zona: archivos con varias hojas planilla → cada hoja = una Zona
  - Mono-zona: archivos con una sola hoja planilla → Zona vacía
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl

# Forzar UTF-8 en stdout/stderr para que los caracteres especiales se impriman
# correctamente en Windows (cp1252 no soporta algunos símbolos).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

# Apunta el path al raíz del proyecto para poder importar los módulos
sys.path.insert(0, str(Path(__file__).parent.parent))

from connectors.airtable import create_record, get_all_records  # noqa: E402

# ---------------------------------------------------------------------------
# Patrones de detección
# ---------------------------------------------------------------------------

# Hojas que se ignoran por nombre
_SKIP_RE = re.compile(
    r"(resumen|pagos|medicion|medición|memo|portada|indice|índice|notas?)",
    re.IGNORECASE,
)

# Cabecera del certificado: "PLANILLA DE CERTIFICADO Nº 3"
_CERT_HDR_RE = re.compile(
    r"PLANILLA\s+DE\s+CERTIFICADO\s+N[ºª°]?\s*(\d+)",
    re.IGNORECASE,
)

# Fecha dentro del mismo string: "FECHA: 30/08/2025" o "FECHA:05/02/2.025"
_FECHA_RE = re.compile(
    r"FECHA[:\s]+(\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2}[\d.]*)",
    re.IGNORECASE,
)

# Fila de totales — detiene el parseo
_TOTAL_RE = re.compile(r"TOTAL\s+GS", re.IGNORECASE)

# Sub-ítem: empieza con letra + guion, ej "a- ", "b- "
_SUB_RE = re.compile(r"^[a-zA-Z]-\s")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _v(row: tuple, idx: int):
    """Valor de celda en posición idx, None si fuera de rango."""
    return row[idx] if idx < len(row) else None


def _blank(val) -> bool:
    return val is None or str(val).strip() == ""


def _float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(",", ".").replace("\xa0", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _normalize_fecha(raw: str) -> str:
    """Convierte '30/08/2025', '05/02/2.025', '30-08-2025' → 'YYYY-MM-DD'."""
    raw = raw.strip().replace(".", "/")
    # Limpiar año con puntos internos como "2.025" → "2025"
    raw = re.sub(r"(\d)\.(\d{3})\b", r"\1\2", raw)
    parts = re.split(r"[/\-]", raw)
    if len(parts) == 3:
        if len(parts[0]) == 4:          # YYYY-MM-DD
            return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
        else:                           # DD-MM-YYYY
            return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
    return raw


# ---------------------------------------------------------------------------
# Detección de layout de columnas
# ---------------------------------------------------------------------------

def _detect_layout(ws, search_from: int) -> tuple[int, ...] | None:
    """Devuelve (i_item, i_rubro, i_unid, i_cant, i_pu, i_ant, i_act)
    detectando si Item está en col A (idx 0) o col B (idx 1).
    Retorna None si no se detecta ningún layout válido.
    """
    for row in ws.iter_rows(min_row=search_from, max_row=search_from + 15, values_only=True):
        a, b = (_v(row, 0), _v(row, 1))
        # Variante A: entero positivo en col A
        if isinstance(a, (int, float)) and a == int(a) and int(a) > 0 and not _blank(b):
            # Item=0 Rubro=1 Unid=2 Cant=3 PU=4 PT=5 Ant=6 Act=7
            return (0, 1, 2, 3, 4, 5, 6, 7)
        # Variante B: entero positivo en col B
        if isinstance(b, (int, float)) and b == int(b) and int(b) > 0 and not _blank(_v(row, 2)):
            # Item=1 Rubro=2 Unid=3 Cant=4 PU=5 PT=6 Ant=7 Act=8
            return (1, 2, 3, 4, 5, 6, 7, 8)
    return None


# ---------------------------------------------------------------------------
# Parseo de una hoja
# ---------------------------------------------------------------------------

def _parse_sheet(ws, zona_name: str) -> tuple[int | None, str, list[dict], list[dict]]:
    """Parsea una hoja planilla.

    Retorna (cert_num, fecha_iso, lineas_presup, lineas_cert).
    lineas_presup: presupuesto base (ítem, rubro, cant, PU, etc.)
    lineas_cert:   {pl_key, cant_actual} — solo filas con Actual > 0
    Retorna (None, '', [], []) si la hoja no es una planilla válida.
    """
    # Buscar cabecera del certificado en filas 5-22
    cert_num  = None
    fecha_iso = ""
    hdr_row   = None

    for row in ws.iter_rows(min_row=5, max_row=22):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            m = _CERT_HDR_RE.search(cell.value)
            if m:
                cert_num = int(m.group(1))
                hdr_row  = cell.row
                fm = _FECHA_RE.search(cell.value)
                if fm:
                    fecha_iso = _normalize_fecha(fm.group(1))
                break
        if cert_num is not None:
            break

    if cert_num is None:
        return None, "", [], []

    # Detectar layout de columnas a partir de hdr_row+1
    layout = _detect_layout(ws, hdr_row + 1)
    if layout is None:
        print(f"    ⚠ No se detectó layout de columnas en hoja '{ws.title}'")
        return None, "", [], []

    i_item, i_rubro, i_unid, i_cant, i_pu, _i_pt, i_ant, i_act = layout
    data_start = hdr_row + 2   # saltar fila de encabezados de columna

    lineas_presup: list[dict] = []
    lineas_cert:   list[dict] = []
    current_grupo = ""
    orden = 1

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not any(c is not None for c in row):
            continue

        rubro_val = _v(row, i_rubro)
        if _blank(rubro_val):
            continue

        rubro_str = str(rubro_val).strip()

        # Detener al llegar a la fila TOTAL
        if _TOTAL_RE.search(rubro_str):
            break

        item_val = _v(row, i_item)
        unid_val = _v(row, i_unid)
        cant_val = _v(row, i_cant)
        pu_val   = _v(row, i_pu)
        act_val  = _v(row, i_act)

        # Fila "SIN COTIZAR" suelta (sin unidad ni ítem)
        if "SIN COTIZAR" in rubro_str.upper() and _blank(item_val) and _blank(unid_val):
            continue

        item_str  = str(item_val).strip() if not _blank(item_val) else ""
        cant_f    = _float(cant_val)
        pu_f      = _float(pu_val)
        act_f     = _float(act_val)

        # Encabezado de grupo: tiene (o no) número de ítem, pero sin unidad ni cantidad
        is_group = _blank(unid_val) and _blank(cant_val) and not _SUB_RE.match(rubro_str)
        if is_group:
            current_grupo = rubro_str.rstrip(":")
            continue

        # Clave única para hacer match entre archivos: zona|item|rubro
        pl_key = f"{zona_name}|{item_str}|{rubro_str}"

        # Detectar SIN COTIZAR en observaciones (a veces en col PT o siguientes)
        obs = ""
        for extra_idx in range(i_act + 1, min(i_act + 4, len(row))):
            extra = str(_v(row, extra_idx) or "")
            if "SIN COTIZAR" in extra.upper():
                obs = "SIN COTIZAR"
                break

        lineas_presup.append({
            "key":            pl_key,
            "Zona":           zona_name,
            "ItemNro":        item_str,
            "GrupoNombre":    current_grupo,
            "Rubro":          rubro_str,
            "Unidad":         str(unid_val).strip() if not _blank(unid_val) else "",
            "Cantidad":       cant_f,
            "PrecioUnitario": pu_f,
            "Observaciones":  obs,
            "Orden":          orden,
        })
        orden += 1

        if act_f != 0.0:
            lineas_cert.append({"pl_key": pl_key, "cant_actual": act_f})

    return cert_num, fecha_iso, lineas_presup, lineas_cert


# ---------------------------------------------------------------------------
# Parseo de un archivo completo
# ---------------------------------------------------------------------------

def _parse_file(path: Path) -> tuple[int | None, str, list[dict], list[dict]]:
    """Parsea un .xlsx completo.
    Retorna (cert_num, fecha_iso, all_presup_lineas, all_cert_lineas).
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  ✗ No se pudo abrir el archivo: {e}")
        return None, "", [], []

    valid_sheets: list[tuple[str, int, str, list, list]] = []

    for name in wb.sheetnames:
        if _SKIP_RE.search(name):
            print(f"    ⊘ Hoja '{name}' ignorada (nombre)")
            continue
        ws = wb[name]
        cn, fi, lp, lc = _parse_sheet(ws, name)
        if cn is None:
            print(f"    ⊘ Hoja '{name}' ignorada (sin cabecera de certificado)")
            continue
        print(f"    ✓ Hoja '{name}': Cert {cn} — {fi} — {len(lp)} ítems — {len(lc)} cert lines")
        valid_sheets.append((name, cn, fi, lp, lc))

    if not valid_sheets:
        return None, "", [], []

    cert_num  = valid_sheets[0][1]
    fecha_iso = valid_sheets[0][2]

    # Si hay una sola hoja planilla → Zona vacía (obra sin zonas)
    # Si hay varias hojas planilla → Zona = nombre de hoja
    is_multi_zona = len(valid_sheets) > 1

    all_presup: list[dict] = []
    all_cert:   list[dict] = []

    for sheet_name, _cn, _fi, lp, lc in valid_sheets:
        zona = sheet_name if is_multi_zona else ""
        # Corregir Zona en las lineas
        for l in lp:
            l["Zona"] = zona
            l["key"]  = f"{zona}|{l['ItemNro']}|{l['Rubro']}"
        for l in lc:
            l["pl_key"] = f"{zona}|{l['pl_key'].split('|',1)[1] if '|' in l['pl_key'] else l['pl_key']}"
        all_presup.extend(lp)
        all_cert.extend(lc)

    return cert_num, fecha_iso, all_presup, all_cert


# ---------------------------------------------------------------------------
# Operaciones Airtable
# ---------------------------------------------------------------------------

def _select_obra(carpeta_nombre: str, obras: list[dict]) -> dict | None:
    """Muestra todas las obras disponibles y pide que el usuario seleccione cuál
    corresponde a la carpeta. Retorna la obra elegida o None para saltear."""
    print(f"\n  Obras disponibles en Airtable:")
    for i, o in enumerate(obras):
        print(f"    {i+1:>3}. {o.get('Clave',''):30s} {o.get('Nombre','')}")
    print(f"    {'S':>3}. Saltear esta carpeta")
    while True:
        raw = input(f"  ¿Cuál obra corresponde a '{carpeta_nombre}'? (número o S): ").strip()
        if raw.upper() == "S":
            return None
        try:
            idx = int(raw) - 1
            if 0 <= idx < len(obras):
                return obras[idx]
        except ValueError:
            pass
        print("  Número inválido, intentá de nuevo.")


def _load_existing_presup(obra_id: str) -> dict[str, str]:
    """Carga las líneas de presupuesto ya existentes en Airtable para esta obra.
    Retorna {pl_key: record_id}. Usado para resumir una migración interrumpida."""
    existing = get_all_records("CertPresupuestoLinea", ["ObraID", "Zona", "ItemNro", "Rubro"])
    result: dict[str, str] = {}
    for r in existing:
        if r.get("ObraID") != obra_id:
            continue
        zona  = r.get("Zona", "") or ""
        item  = r.get("ItemNro", "") or ""
        rubro = r.get("Rubro", "") or ""
        result[f"{zona}|{item}|{rubro}"] = r["_id"]
    return result


def _load_existing_cert_refs(obra_id: str) -> set[str]:
    """Retorna el conjunto de CertRef ya existentes en Airtable para esta obra."""
    existing = get_all_records("CertCabecera", ["ObraID", "CertRef"])
    return {r.get("CertRef", "") for r in existing if r.get("ObraID") == obra_id}


def _upload_presupuesto(obra_id: str, lineas: list[dict], dry_run: bool) -> dict[str, str]:
    """Sube las líneas de presupuesto. Retorna {pl_key: record_id}.
    Si ya existen líneas para esta obra en Airtable, las reutiliza sin crear duplicados."""
    if dry_run:
        return {l["key"]: f"dry_pres_{i}" for i, l in enumerate(lineas, 1)}

    # Cargar las que ya existen (por si la migración fue interrumpida)
    key_to_id = _load_existing_presup(obra_id)
    if key_to_id:
        print(f"    {len(key_to_id)} líneas ya existentes en Airtable, reutilizando...")

    nuevas = [l for l in lineas if l["key"] not in key_to_id]
    for i, l in enumerate(nuevas, 1):
        fields = {
            "ObraID":         [obra_id],
            "Orden":          l["Orden"],
            "Zona":           l["Zona"],
            "ItemNro":        l["ItemNro"],
            "GrupoNombre":    l["GrupoNombre"],
            "Rubro":          l["Rubro"],
            "Unidad":         l["Unidad"],
            "Cantidad":       l["Cantidad"] or None,
            "PrecioUnitario": l["PrecioUnitario"] or None,
            "Observaciones":  l["Observaciones"],
            "SinCotizar":     l["Observaciones"].upper() == "SIN COTIZAR" if l.get("Observaciones") else False,
        }
        rec_id = create_record("CertPresupuestoLinea", fields)
        key_to_id[l["key"]] = rec_id
        if i % 10 == 0 or i == len(nuevas):
            print(f"    {i}/{len(nuevas)} líneas nuevas...", end="\r")
    if nuevas:
        print()
    return key_to_id


def _upload_cert(
    obra_id:   str,
    obra_clave: str,
    cert_num:  int,
    fecha_iso: str,
    lineas_cert: list[dict],
    key_to_id: dict[str, str],
    dry_run:   bool,
    existing_refs: set[str] | None = None,
):
    """Sube una CertCabecera + sus CertLinea. Saltea si CertRef ya existe."""
    cert_ref = f"{obra_clave}-CERT-{cert_num}"

    if not dry_run and existing_refs is not None and cert_ref in existing_refs:
        print(f"    [SKIP] {cert_ref} ya existe en Airtable")
        return

    cab_fields = {
        "CertRef":          cert_ref,
        "ObraID":           [obra_id],
        "Numero":           cert_num,
        "FechaCertificado": fecha_iso or None,
        "Estado":           "Confirmado",
        "Observaciones":    "Importado desde Excel histórico",
    }

    if dry_run:
        cab_id = f"dry_cab_{cert_num}"
        print(f"    [DRY] Cabecera: {cert_ref} — {fecha_iso}")
    else:
        cab_id = create_record("CertCabecera", cab_fields)

    uploaded      = 0
    missing_keys  = []

    for lc in lineas_cert:
        if lc["cant_actual"] == 0.0:
            continue
        pl_id = key_to_id.get(lc["pl_key"])
        if not pl_id:
            missing_keys.append(lc["pl_key"])
            continue
        line_fields = {
            "CabeceraID":          [cab_id],
            "PresupuestoLineaID":  [pl_id],
            "CantidadCertificada": lc["cant_actual"],
            "LineaRef":            f"{cert_ref}-L{uploaded+1}",
        }
        if not dry_run:
            create_record("CertLinea", line_fields)
        uploaded += 1

    print(f"    ✓ {uploaded} líneas certificadas subidas")
    if missing_keys:
        print(f"    ⚠ {len(missing_keys)} ítems sin match en el presupuesto:")
        for k in missing_keys[:5]:
            print(f"      {k}")
        if len(missing_keys) > 5:
            print(f"      ... y {len(missing_keys)-5} más")


# ---------------------------------------------------------------------------
# Procesamiento por carpeta
# ---------------------------------------------------------------------------

def _process_obra_folder(carpeta: Path, obra: dict, dry_run: bool) -> bool:
    """Procesa todos los .xlsx de una carpeta para una obra dada.
    Retorna True si se procesó algo, False si no había archivos válidos.
    """
    obra_id    = obra["_id"]
    obra_clave = obra.get("Clave", obra.get("Nombre", "?"))

    # ── Parsear archivos ──────────────────────────────────────────────────────
    xlsx_files = sorted(carpeta.glob("*.xlsx"))
    if not xlsx_files:
        print(f"  ✗ No hay archivos .xlsx en {carpeta}")
        return False

    print(f"  {len(xlsx_files)} archivo(s) encontrado(s)\n")

    parsed: list[tuple[int, str, list, list, str]] = []
    for f in xlsx_files:
        print(f"  -> {f.name}")
        cert_num, fecha_iso, presup, cert_lines = _parse_file(f)
        if cert_num is None:
            print(f"     [SALTADO]\n")
            continue
        parsed.append((cert_num, fecha_iso, presup, cert_lines, f.name))
        print(f"     Cert {cert_num} -- {fecha_iso} -- {len(presup)} items -- {len(cert_lines)} lineas\n")

    if not parsed:
        print("  [ERROR] Ningun archivo parseado correctamente.")
        return False

    # Ordenar por número de certificado (orden aproximado del Excel) y luego
    # reasignar números secuenciales 1, 2, 3... para evitar duplicados.
    parsed.sort(key=lambda x: x[0])
    parsed = [
        (seq, fecha_iso, presup, cert_lines, fname)
        for seq, (_, fecha_iso, presup, cert_lines, fname) in enumerate(parsed, start=1)
    ]
    print(f"  Numeracion secuencial asignada: 1 a {len(parsed)}")

    # -- Construir presupuesto unificado (union de todos los archivos) ---------
    # Itera los archivos en orden de cert_num. Cada item nuevo (por key) se agrega
    # al final del presupuesto maestro con Orden incremental. Esto preserva los
    # items originales en su orden y agrega los adicionales en orden de aparicion.
    merged_presup: list[dict] = []
    seen_keys: set[str] = set()
    next_orden = 1
    for _, _, presup, _, fname in parsed:
        nuevos = 0
        for item in presup:
            if item["key"] not in seen_keys:
                item = dict(item)           # copia para no mutar el original
                item["Orden"] = next_orden
                merged_presup.append(item)
                seen_keys.add(item["key"])
                next_orden += 1
                nuevos += 1
        if nuevos:
            print(f"  +{nuevos} items nuevos desde '{fname}'")

    print(f"  Subiendo presupuesto unificado ({len(merged_presup)} lineas)...")
    key_to_id = _upload_presupuesto(obra_id, merged_presup, dry_run)
    print(f"     {len(key_to_id)} lineas {'simuladas' if dry_run else 'subidas'}\n")

    # -- Subir certificados ---------------------------------------------------
    existing_refs = None if dry_run else _load_existing_cert_refs(obra_id)
    if existing_refs:
        print(f"  {len(existing_refs)} certificado(s) ya existentes en Airtable, se saltearán.")

    print(f"  Subiendo {len(parsed)} certificado(s)...\n")
    for cert_num, fecha_iso, _, cert_lines, fname in parsed:
        print(f"  -> Cert {cert_num} ({fname})")
        _upload_cert(obra_id, obra_clave, cert_num, fecha_iso, cert_lines, key_to_id, dry_run, existing_refs)
        print()

    status = "simulada (dry-run)" if dry_run else "completada"
    print(f"  [OK] Migracion {status}: {obra_clave} -- {len(merged_presup)} pres. + {len(parsed)} certs.")
    return True


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Migrar certificados históricos desde Excel a Airtable"
    )
    parser.add_argument(
        "--raiz", required=True,
        help="Carpeta raíz que contiene subcarpetas (una por obra) con los .xlsx",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Simula la importación sin escribir en Airtable",
    )
    args = parser.parse_args()

    raiz = Path(args.raiz)
    if not raiz.is_dir():
        print(f"✗ Carpeta raíz no encontrada: {raiz}")
        sys.exit(1)

    subcarpetas = sorted([p for p in raiz.iterdir() if p.is_dir()])
    if not subcarpetas:
        print(f"✗ No hay subcarpetas en {raiz}")
        sys.exit(1)

    # Cargar obras de Airtable una sola vez
    print("Cargando obras desde Airtable...")
    raw_obras = get_all_records("obras", ["Nombre", "Clave"])
    obras = [
        {
            "_id":    r["_id"],
            "Nombre": r.get("Nombre", ""),
            "Clave":  r.get("Clave", ""),
        }
        for r in raw_obras
    ]
    print(f"   {len(obras)} obras cargadas\n")

    if args.dry_run:
        print("  *** DRY RUN -- no se escribira nada en Airtable ***\n")

    # Procesar cada subcarpeta
    resultados: list[tuple[str, str]] = []  # (carpeta, resultado)
    for carpeta in subcarpetas:
        print(f"\n{'='*60}")
        print(f"Carpeta: {carpeta.name}")
        print(f"{'='*60}")
        obra = _select_obra(carpeta.name, obras)
        if obra is None:
            print(f"  [SALTADA]\n")
            resultados.append((carpeta.name, "saltada"))
            continue
        ok = _process_obra_folder(carpeta, obra, args.dry_run)
        resultados.append((carpeta.name, "ok" if ok else "sin archivos"))

    # Resumen final
    print(f"\n{'='*60}")
    print("Resumen:")
    for carpeta_nombre, estado in resultados:
        icon = "[OK]" if estado == "ok" else ("[SALTADA]" if estado == "saltada" else "[SIN ARCHIVOS]")
        print(f"  {icon} {carpeta_nombre}: {estado}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
